from __future__ import annotations

import hashlib
import io
import os
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from caos.artifacts.loan_universe import (
    HEADERS,
    LoanWorkbookValidationError,
    parse_loan_workbook,
    source_bytes,
)
from caos.config import Settings
from caos.http import create_app
from caos.memory_ledgers import MemoryLedgerSet
from caos.postgres_ledgers import PostgresLedgerSet


def test_source_bytes_rejects_noncanonical_digest_before_path_lookup(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError, match="RV_SOURCE_BYTES_UNAVAILABLE"):
        source_bytes({"sha256": "../../outside"}, tmp_path)


def _row(
    *,
    borrower: str = "Access CIG LLC",
    bloomberg: str = "BLS202439",
    figi: str = "BBG01WMCP303",
    margin: object = 400,
    maturity: object = date(2030, 8, 19),
    change_1d: object = 0.5,
) -> list[object]:
    return [
        "Access CIG",
        borrower,
        "Provides records management and digital information services.",
        "Business Services",
        "Records Management",
        "Private",
        bloomberg,
        figi,
        "B1",
        "1L Gtd. Sr. Secd",
        "B3 / B",
        1475,
        margin,
        maturity,
        88,
        90,
        change_1d,
        0.5,
        1,
        -2,
        -4.13,
        1,
        -7.5,
        11.2,
        851,
    ]


def _sheet(workbook: Workbook, title: str, rows: list[list[object]], *, workbook_date: object = date(2026, 8, 24)) -> None:
    sheet = workbook.create_sheet(title)
    sheet["B1"] = "Date"
    sheet["B2"] = workbook_date
    for column, header in enumerate(HEADERS, start=1):
        sheet.cell(row=5, column=column, value=header)
    for row_number, values in enumerate(rows, start=6):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_number, column=column, value=value)
    summary_row = 6 + len(rows) + 1
    sheet.cell(row=summary_row, column=1, value="Index Statistics")
    sheet.cell(row=summary_row, column=2, value="US Leveraged Loan Index")


def _workbook_bytes(
    *,
    first_rows: list[list[object]] | None = None,
    second_rows: list[list[object]] | None = None,
    second_date: object = date(2026, 8, 24),
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", first_rows or [_row()])
    _sheet(
        workbook,
        "Healthcare IT",
        second_rows or [_row(borrower="FinThrive Inc", bloomberg="BLS1989347", figi="BBG01THRIVE1")],
        workbook_date=second_date,
    )
    _sheet(workbook, "Hidden Support", [_row(borrower="Hidden", bloomberg="HIDDEN", figi="BBG00HIDDEN1")])
    workbook["Hidden Support"].sheet_state = "hidden"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _codes(error: LoanWorkbookValidationError) -> set[str]:
    return {finding["code"] for finding in error.findings}


def _import_record(
    case_id: str,
    parsed: dict[str, object],
    *,
    status: str = "ACTIVE",
    findings: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    return {
        "case_id": case_id,
        "source_id": parsed["source_id"],
        "source_filename": "REF_CP-3_Sector_RV.xlsx",
        "source_sha256": parsed["source_sha256"],
        "workbook_date": parsed.get("workbook_date"),
        "template_version": parsed["template_version"],
        "importer_version": parsed["importer_version"],
        "universe_digest": parsed.get("universe_digest"),
        "row_count": parsed.get("row_count", 0),
        "status": status,
        "findings": findings or [],
    }


def _source_record(case_id: str, sha256: str) -> dict[str, object]:
    return {
        "case_id": case_id,
        "filename": "REF_CP-3_Sector_RV.xlsx",
        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "sha256": sha256,
        "vault_path": "/tmp/test-loan-source.xlsx",
        "bytes": 1,
        "blocks": [],
    }


def test_cp3_workbook_maps_all_visible_sector_rows_with_source_units() -> None:
    content = _workbook_bytes(
        first_rows=[_row(margin=400, change_1d=0.5), _row(borrower="Apex Group", bloomberg="BLS5005287", figi="BBG01S807689", margin="#N/A", change_1d="N/A")]
    )

    parsed = parse_loan_workbook(content, source_id="src_1", source_sha256="a" * 64)

    assert parsed["workbook_date"] == "2026-08-24"
    assert parsed["row_count"] == 3
    assert [row["sector"] for row in parsed["rows"]] == ["Healthcare IT", "IT Services", "IT Services"]
    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["instrument_key"] == "FIGI:BBG01WMCP303"
    assert access["size_mn"] == 1475.0
    assert access["margin_bps"] == 400.0
    assert access["bid_points"] == 88.0
    assert access["change_3m_points"] == -2.0
    assert access["mid_ytm_pct"] == 11.2
    assert access["mid_3y_dm_bps"] == 851.0
    assert access["maturity_date"] == "2030-08-19"
    assert access["source_locators"] == [{"sheet": "IT Services", "row": 6}]
    apex = next(row for row in parsed["rows"] if row["borrower_name"] == "Apex Group")
    assert apex["margin_bps"] is None and apex["change_1d_points"] is None
    assert all(row["borrower_name"] != "Hidden" for row in parsed["rows"])
    assert parsed["universe_digest"] == parse_loan_workbook(content, source_id="src_1", source_sha256="a" * 64)["universe_digest"]


def test_duplicate_rows_collapse_and_preserve_every_locator() -> None:
    duplicate = _row()
    content = _workbook_bytes(first_rows=[duplicate, duplicate])

    parsed = parse_loan_workbook(content, source_id="src_1", source_sha256="b" * 64)

    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["source_locators"] == [{"sheet": "IT Services", "row": 6}, {"sheet": "IT Services", "row": 7}]
    assert parsed["row_count"] == 2


def test_blank_optional_cells_remain_null_and_keep_their_column_locator() -> None:
    parsed = parse_loan_workbook(
        _workbook_bytes(first_rows=[_row(margin=None)]),
        source_id="src_1",
        source_sha256="0" * 64,
    )

    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["margin_bps"] is None


def test_bloomberg_alias_reconciles_a_missing_figi_before_duplicate_collapse() -> None:
    parsed = parse_loan_workbook(
        _workbook_bytes(
            first_rows=[
                _row(bloomberg="LOAN1", figi="#N/A"),
                _row(bloomberg="LOAN1", figi="FIGI1"),
            ]
        ),
        source_id="src_1",
        source_sha256="9" * 64,
    )

    access = next(row for row in parsed["rows"] if row["bloomberg_loan_id"] == "LOAN1")
    assert access["instrument_key"] == "FIGI:FIGI1"
    assert access["figi"] == "FIGI1"
    assert access["source_locators"] == [
        {"sheet": "IT Services", "row": 6},
        {"sheet": "IT Services", "row": 7},
    ]
    assert parsed["row_count"] == 2


def test_conflicting_duplicates_and_identifier_mappings_reject_the_candidate() -> None:
    conflict = _row(margin=425)
    mapping_conflict = _row(borrower="Other", bloomberg="OTHER", figi="BBG01WMCP303")
    with pytest.raises(LoanWorkbookValidationError) as raised:
        parse_loan_workbook(_workbook_bytes(first_rows=[_row(), conflict, mapping_conflict]), source_id="src", source_sha256="c" * 64)

    assert {"RV_DUPLICATE_CONFLICT", "RV_ID_CONFLICT"}.issubset(_codes(raised.value))


def test_partial_headers_and_conflicting_dates_reject_the_complete_workbook() -> None:
    content = _workbook_bytes(second_date="25/08/2026")
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", [_row()])
    workbook["IT Services"].cell(row=5, column=13, value="Coupon")
    partial = io.BytesIO()
    workbook.save(partial)
    workbook.close()

    with pytest.raises(LoanWorkbookValidationError) as date_error:
        parse_loan_workbook(content, source_id="src", source_sha256="d" * 64)
    with pytest.raises(LoanWorkbookValidationError) as header_error:
        parse_loan_workbook(partial.getvalue(), source_id="src", source_sha256="e" * 64)

    assert "RV_WORKBOOK_DATE_CONFLICT" in _codes(date_error.value)
    assert {"RV_TEMPLATE_PARTIAL", "RV_TEMPLATE_MISSING"}.issubset(_codes(header_error.value))


@pytest.mark.parametrize(
    ("row", "expected_code"),
    [
        (_row(borrower="", bloomberg="", figi=""), "RV_BORROWER_MISSING"),
        (_row(margin="not-a-number"), "RV_NUMBER_INVALID"),
        (_row(margin="1e999"), "RV_NUMBER_NON_FINITE"),
        (_row(maturity="31/31/2030"), "RV_MATURITY_INVALID"),
        (_row(maturity=0), "RV_MATURITY_INVALID"),
    ],
)
def test_invalid_rows_fail_closed(row: list[object], expected_code: str) -> None:
    with pytest.raises(LoanWorkbookValidationError) as raised:
        parse_loan_workbook(_workbook_bytes(first_rows=[row]), source_id="src", source_sha256="f" * 64)

    assert expected_code in _codes(raised.value)


def test_formula_without_cached_value_becomes_null_without_execution() -> None:
    formula_row = _row(margin="=200+200")

    parsed = parse_loan_workbook(_workbook_bytes(first_rows=[formula_row]), source_id="src", source_sha256="1" * 64)

    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["margin_bps"] is None


def test_unsafe_package_parts_reject_before_workbook_parsing() -> None:
    content = io.BytesIO(_workbook_bytes())
    with zipfile.ZipFile(content, "a") as archive:
        archive.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")

    with pytest.raises(LoanWorkbookValidationError) as raised:
        parse_loan_workbook(content.getvalue(), source_id="src", source_sha256="2" * 64)

    assert "RV_PACKAGE_ACTIVE_CONTENT" in _codes(raised.value)


def test_external_relationship_with_xml_whitespace_is_rejected() -> None:
    content = io.BytesIO(_workbook_bytes())
    with zipfile.ZipFile(content, "a") as archive:
        archive.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
            <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
              <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="https://example.invalid" TargetMode = "External"/>
            </Relationships>""",
        )

    with pytest.raises(LoanWorkbookValidationError) as raised:
        parse_loan_workbook(content.getvalue(), source_id="src", source_sha256="8" * 64)

    assert "RV_PACKAGE_EXTERNAL_LINK" in _codes(raised.value)


def test_workbook_sheet_limit_rejects_without_scanning_beyond_the_cap() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", [_row()])
    for index in range(64):
        sheet = workbook.create_sheet(f"Hidden {index:02d}")
        sheet.sheet_state = "hidden"
    content = io.BytesIO()
    workbook.save(content)
    workbook.close()

    with pytest.raises(LoanWorkbookValidationError) as raised:
        parse_loan_workbook(content.getvalue(), source_id="src", source_sha256="3" * 64)

    assert "RV_WORKSHEET_LIMIT" in _codes(raised.value)



def test_loan_universe_migration_has_atomic_identity_and_active_constraints() -> None:
    migration = (Path(__file__).parents[1] / "server" / "migrations" / "003_rv_loan_universes.sql").read_text()

    assert "UNIQUE (case_id, source_sha256, template_version, importer_version)" in migration
    assert "WHERE status = 'ACTIVE'" in migration
    assert "PRIMARY KEY (universe_id, instrument_key)" in migration


def test_loan_universe_source_foreign_key_migration_backfills_before_validation() -> None:
    migration = (Path(__file__).parents[1] / "server" / "migrations" / "004_rv_loan_universe_source_fk.sql").read_text()
    migrations = (Path(__file__).parents[1] / "server" / "migrations").glob("*.sql")

    assert "JOIN rv_loan_universes ON rv_loan_universes.source_id = source.key" in migration
    assert "FOREIGN KEY (source_id) REFERENCES sources(id) ON DELETE RESTRICT NOT VALID" in migration
    assert "VALIDATE CONSTRAINT rv_loan_universes_source_id_fkey" in migration
    assert sum(
        "FOREIGN KEY (source_id) REFERENCES sources(id)" in path.read_text()
        for path in migrations
    ) == 1


def test_postgres_loan_import_converges_and_normalizes_rows() -> None:
    database_url = os.getenv("CAOS_TEST_DATABASE_URL")
    if not database_url:
        pytest.skip("CAOS_TEST_DATABASE_URL is required for durable loan-universe proof")
    import psycopg

    content = _workbook_bytes()
    first = PostgresLedgerSet(database_url)
    case = first.runs.create_case("Loan RV", "Issuer", "Services", "analyst")
    source = first.sources.ingest(
        _source_record(case["id"], hashlib.sha256(content).hexdigest()), "analyst"
    )
    second = PostgresLedgerSet(database_url)
    parsed = parse_loan_workbook(
        content, source_id=source["id"], source_sha256=source["sha256"]
    )
    first_record = _import_record(case["id"], parsed)
    second_record = _import_record(case["id"], parsed)

    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(
            executor.map(
                lambda item: item[0].sources.save_loan_universe_import(
                    item[1], parsed["rows"], "analyst"
                ),
                ((first, first_record), (second, second_record)),
            )
        )

    assert sum(created for _record, created in results) == 1
    universe_id = results[0][0]["id"]
    assert results[1][0]["id"] == universe_id
    check = PostgresLedgerSet(database_url)
    active = check.sources.active_loan_universe(case["id"])
    assert active is not None and active["id"] == universe_id and len(active["rows"]) == parsed["row_count"]
    with psycopg.connect(database_url.replace("postgresql+psycopg://", "postgresql://")) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT universe.status, (SELECT count(*) FROM rv_loan_rows WHERE universe_id=%s), source.id "
                "FROM rv_loan_universes AS universe JOIN sources AS source ON source.id=universe.source_id "
                "WHERE universe.id=%s",
                (universe_id, universe_id),
            )
            row = cursor.fetchone()
    assert row == ("ACTIVE", parsed["row_count"], source["id"])


def test_postgres_import_and_source_withdrawal_converge_without_active_residue() -> None:
    database_url = os.getenv("CAOS_TEST_DATABASE_URL")
    if not database_url:
        pytest.skip("CAOS_TEST_DATABASE_URL is required for durable loan-universe proof")
    content = _workbook_bytes()
    importer = PostgresLedgerSet(database_url)
    case = importer.runs.create_case(
        "Loan withdrawal race", "Issuer", "Services", "analyst"
    )
    source = importer.sources.ingest(
        _source_record(case["id"], hashlib.sha256(content).hexdigest()), "analyst"
    )
    withdrawer = PostgresLedgerSet(database_url)
    parsed = parse_loan_workbook(
        content,
        source_id=source["id"],
        source_sha256=source["sha256"],
    )
    record = _import_record(case["id"], parsed)

    def activate() -> str:
        try:
            importer.sources.save_loan_universe_import(
                record, parsed["rows"], "analyst"
            )
            return "activated"
        except ValueError as exc:
            assert str(exc) == "RV_SOURCE_NOT_ACTIVE"
            return "withdrawn-first"

    with ThreadPoolExecutor(max_workers=2) as executor:
        activation = executor.submit(activate)
        withdrawal = executor.submit(
            withdrawer.sources.withdraw, case["id"], source["id"], "analyst"
        )
        assert activation.result() in {"activated", "withdrawn-first"}
        assert withdrawal.result() is not None

    check = PostgresLedgerSet(database_url)
    assert check.sources.get_source(source["id"])["withdrawn"] is True
    assert check.sources.active_loan_universe(case["id"]) is None


def _client(
    tmp_path: Path, ledger_set: MemoryLedgerSet | None = None
) -> TestClient:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1] / "server" / "caos" / "methodology" / "vendor" / "deploy_v",
    )
    return TestClient(create_app(settings, ledger_set or MemoryLedgerSet()))


def _upload_workbook(client: TestClient, case_id: str, content: bytes, name: str = "REF_CP-3_Sector_RV.xlsx") -> dict[str, object]:
    response = client.post(
        f"/api/cases/{case_id}/sources",
        files={"file": (name, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201
    return response.json()


def test_case_api_imports_reads_and_idempotently_returns_the_active_universe(tmp_path: Path) -> None:
    with _client(tmp_path) as client:
        case_id = client.post("/api/cases", json={"name": "Loan RV", "issuer": "Issuer", "sector": "Services"}).json()["id"]
        source = _upload_workbook(client, case_id, _workbook_bytes())

        imported = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]})
        repeated = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]})
        active = client.get(f"/api/cases/{case_id}/rv/loan-universes/active")

        assert imported.status_code == 201 and repeated.status_code == 200
        assert repeated.json()["id"] == imported.json()["id"]
        assert active.status_code == 200 and active.json()["status"] == "ACTIVE"
        assert active.json()["universe"]["id"] == imported.json()["id"]
        assert active.json()["universe"]["source_id"] == source["id"]
        assert len(active.json()["rows"]) == 2
        assert active.json()["rows"][0]["mid_3y_dm_bps"] == 851.0


def test_invalid_import_returns_structured_findings_and_preserves_prior_active_universe(tmp_path: Path) -> None:
    with _client(tmp_path) as client:
        case_id = client.post("/api/cases", json={"name": "Loan RV", "issuer": "Issuer", "sector": "Services"}).json()["id"]
        valid_source = _upload_workbook(client, case_id, _workbook_bytes())
        active_id = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": valid_source["id"]}).json()["id"]

        workbook = Workbook()
        workbook.remove(workbook.active)
        _sheet(workbook, "IT Services", [_row()])
        workbook["IT Services"].cell(row=5, column=13, value="Coupon")
        invalid_bytes = io.BytesIO()
        workbook.save(invalid_bytes)
        workbook.close()
        invalid_source = _upload_workbook(client, case_id, invalid_bytes.getvalue(), "changed-template.xlsx")

        rejected = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": invalid_source["id"]})
        repeated = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": invalid_source["id"]})
        active = client.get(f"/api/cases/{case_id}/rv/loan-universes/active").json()

        assert rejected.status_code == repeated.status_code == 422
        assert rejected.json()["detail"]["code"] == "RV_WORKBOOK_INVALID"
        assert {finding["code"] for finding in rejected.json()["detail"]["findings"]} >= {"RV_TEMPLATE_PARTIAL", "RV_TEMPLATE_MISSING"}
        assert repeated.json()["detail"]["universe_id"] == rejected.json()["detail"]["universe_id"]
    assert active["universe"]["id"] == active_id


def test_import_rejects_vault_bytes_that_do_not_match_the_source_digest(tmp_path: Path) -> None:
    ledger_set = MemoryLedgerSet()
    with _client(tmp_path, ledger_set) as client:
        case_id = client.post("/api/cases", json={"name": "Loan RV", "issuer": "Issuer", "sector": "Services"}).json()["id"]
        source = _upload_workbook(client, case_id, _workbook_bytes())
        source_digest = str(source["sha256"])
        vault_path = tmp_path / "vault" / "sources" / source_digest[:2] / source_digest
        vault_path.write_bytes(_workbook_bytes(first_rows=[_row(margin=425)]))

        rejected = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]})

        assert rejected.status_code == 409
        assert rejected.json()["detail"]["code"] == "RV_SOURCE_INTEGRITY_MISMATCH"
        assert ledger_set.sources.active_loan_universe(case_id) is None


def test_loan_universe_api_is_case_scoped_and_reader_safe(tmp_path: Path) -> None:
    ledger_set = MemoryLedgerSet()
    with _client(tmp_path, ledger_set) as client:
        case_id = client.post("/api/cases", json={"name": "Loan RV", "issuer": "Issuer", "sector": "Services"}).json()["id"]
        other_case_id = client.post("/api/cases", json={"name": "Other", "issuer": "Other", "sector": "Services"}).json()["id"]
        source = _upload_workbook(client, case_id, _workbook_bytes())
        assert client.post(f"/api/cases/{other_case_id}/rv/loan-universes", json={"source_id": source["id"]}).status_code == 404

        assert ledger_set.runs.add_member(case_id, "local-analyst", "reader", "READER", "ADMIN")
        reader_headers = {"x-forwarded-user": "reader", "x-caos-role": "READER"}
        assert client.get(f"/api/cases/{case_id}/rv/loan-universes/active", headers=reader_headers).status_code == 200
        assert client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]}, headers=reader_headers).status_code == 403
        assert client.get(f"/api/cases/{case_id}/rv/loan-universes/active", headers={"x-forwarded-user": "outsider"}).status_code == 404


def test_source_withdrawal_deactivates_loan_universe(tmp_path: Path) -> None:
    ledger_set = MemoryLedgerSet()
    with _client(tmp_path, ledger_set) as client:
        case_id = client.post("/api/cases", json={"name": "Loan RV", "issuer": "Issuer", "sector": "Services"}).json()["id"]
        source = _upload_workbook(client, case_id, _workbook_bytes())
        client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]})
        assert client.post(f"/api/cases/{case_id}/sources/{source['id']}/withdraw").status_code == 200
        assert client.get(f"/api/cases/{case_id}/rv/loan-universes/active").json() == {
            "status": "NO_ACTIVE_UNIVERSE",
            "universe": None,
            "rows": [],
        }


def test_cp3_artifact_binds_the_pinned_normalized_loan_universe(tmp_path: Path) -> None:
    with _client(tmp_path) as client:
        case_id = client.post("/api/cases", json={"name": "Loan RV", "issuer": "Issuer", "sector": "Services"}).json()["id"]
        source = _upload_workbook(client, case_id, _workbook_bytes())
        universe = client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]}).json()
        active_rows = client.get(f"/api/cases/{case_id}/rv/loan-universes/active").json()["rows"]
        started = client.post(f"/api/cases/{case_id}/runs", json={"pathway": "RELATIVE_VALUE", "depth": "full"})
        assert started.status_code == 202
        for _ in range(120):
            run = client.get(f"/api/runs/{started.json()['id']}").json()
            if run["status"] in {"succeeded", "failed"}:
                break
            time.sleep(0.05)
        assert run["status"] == "succeeded"
        cp3_node = next(node for node in run["nodes"] if node["module_id"] == "CP-3")
        cp3 = client.app.state.ledger_set.runs.get_artifact(cp3_node["artifact_id"])
        identity = cp3["payload"]["lineage"]["loan_universe"]
        assert identity["id"] == universe["id"]
        assert identity["universe_digest"] == universe["universe_digest"]
        assert identity["source_id"] == source["id"]
        assert cp3["payload"]["provenance"]["loan_universe"] == identity
        assert cp3["payload"]["inputs"]["loan_universe"] == {"identity": identity, "rows": active_rows}
        assert {row["source_locators"][0]["sheet"] for row in active_rows} == {"IT Services", "Healthcare IT"}
        assert all(row["source_locators"][0]["row"] == 6 for row in active_rows)
