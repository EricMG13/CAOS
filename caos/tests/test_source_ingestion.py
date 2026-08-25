from __future__ import annotations

import asyncio
import io
import re
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook
from pypdf import PdfWriter
import pytest
from starlette.datastructures import UploadFile

from caos.config import Settings
from caos.http import create_app
from caos.memory_ledgers import MemoryLedgerSet
from caos.sources import domain as source_domain
from caos.sources.domain import Vault, ingest_upload


def test_zero_byte_source_is_rejected_before_source_set_creation(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Empty source", "issuer": "Issuer", "sector": "Test"},
        ).json()["id"]

        upload = client.post(
            f"/api/cases/{case_id}/sources",
            files={"file": ("empty.txt", b"", "text/plain")},
        )

        assert upload.status_code == 422
        assert upload.json()["detail"] == "source is empty"
        assert (
            client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
            == "NEEDS_SOURCE"
        )


def test_source_ingestion_rolls_back_metadata_when_persistence_fails(
    tmp_path: Path,
) -> None:
    class FailingCatalog:
        def __init__(self, catalog: object) -> None:
            self.catalog = catalog
            self.fail_ingest = False

        def ingest(self, source: dict[str, object], actor: str) -> dict[str, object]:
            if self.fail_ingest:
                raise RuntimeError("database unavailable")
            return self.catalog.ingest(source, actor)  # type: ignore[attr-defined,no-any-return]

    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )

    def upload(
        catalog: FailingCatalog, case_id: str, name: str = "source.txt"
    ) -> dict[str, object]:
        return asyncio.run(
            ingest_upload(
                catalog,
                Vault(settings),
                case_id,
                "analyst",
                UploadFile(file=io.BytesIO(b"Revenue 100"), filename=name),
                settings.max_upload_bytes,
            )
        )

    ledgers = MemoryLedgerSet()
    catalog = FailingCatalog(ledgers.sources)
    case_id = ledgers.runs.create_case("Source", "Issuer", "Sector", "analyst")["id"]
    catalog.fail_ingest = True
    with pytest.raises(RuntimeError, match="database unavailable"):
        upload(catalog, case_id)
    assert ledgers.sources.list_sources(case_id) == []
    assert ledgers.sources.current_source_set(case_id) is None

    catalog.fail_ingest = False
    assert upload(catalog, case_id, "recovered.txt")["source_set"]["version"] == 1


def test_source_ingestion_commits_source_and_source_set_together(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    ledgers = MemoryLedgerSet()
    case_id = ledgers.runs.create_case("Source audit", "Issuer", "Sector", "analyst")[
        "id"
    ]

    source = asyncio.run(
        ingest_upload(
            ledgers.sources,
            Vault(settings),
            case_id,
            "analyst",
            UploadFile(file=io.BytesIO(b"Revenue 100"), filename="source.txt"),
            settings.max_upload_bytes,
        )
    )

    assert [item["id"] for item in ledgers.sources.list_sources(case_id)] == [
        source["id"]
    ]
    assert ledgers.sources.current_source_set(case_id) == source["source_set"]


def test_evidence_free_text_csv_and_xlsx_are_rejected_before_source_set_creation(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    workbook = Workbook()
    blank_xlsx = io.BytesIO()
    workbook.save(blank_xlsx)
    uploads = [
        ("spaces.txt", b" \n\t\n", "text/plain"),
        ("spaces.csv", b",,\n,,\n", "text/csv"),
        ("quoted-empty.csv", b',"",\n', "text/csv"),
        (
            "blank.xlsx",
            blank_xlsx.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    ]

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        for filename, content, media_type in uploads:
            case_id = client.post(
                "/api/cases",
                json={"name": filename, "issuer": "Issuer", "sector": "Test"},
            ).json()["id"]
            upload = client.post(
                f"/api/cases/{case_id}/sources",
                files={"file": (filename, content, media_type)},
            )

            assert upload.status_code == 422
            assert upload.json()["detail"] == "source contains no extractable evidence"
            assert (
                client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
                == "NEEDS_SOURCE"
            )


def test_xlsx_evidence_extraction_rejects_truncated_workbooks(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "first"
    sheet["A2"] = "second"
    sheet["A3"] = "row-limit-sentinel"
    content = io.BytesIO()
    workbook.save(content)
    workbook.close()
    monkeypatch.setattr(source_domain, "MAX_XLSX_EXTRACT_ROWS", 2)

    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Bounded source", "issuer": "Issuer", "sector": "Test"},
        ).json()["id"]
        upload = client.post(
            f"/api/cases/{case_id}/sources",
            files={
                "file": (
                    "bounded.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert upload.status_code == 422
        assert upload.json()["detail"] == source_domain.EXTRACTION_LIMIT_DETAIL
        assert (
            client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
            == "NEEDS_SOURCE"
        )


def test_source_extraction_rejects_hidden_columns_sheets_and_long_lines(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def workbook_bytes(*, sheets: int = 1, row: int = 1, column: int = 1) -> bytes:
        workbook = Workbook()
        workbook.active.cell(row=row, column=column, value="evidence")
        for index in range(1, sheets):
            workbook.create_sheet(f"Sheet {index + 1}")["A1"] = "evidence"
        content = io.BytesIO()
        workbook.save(content)
        workbook.close()
        return content.getvalue()

    def declare_only_a1(content: bytes) -> bytes:
        source = ZipFile(io.BytesIO(content))
        target_bytes = io.BytesIO()
        with source, ZipFile(target_bytes, "w", ZIP_DEFLATED) as target:
            for entry in source.infolist():
                data = source.read(entry.filename)
                if entry.filename == "xl/worksheets/sheet1.xml":
                    data = re.sub(
                        rb'<dimension ref="[^"]+"\s*/>',
                        b'<dimension ref="A1:A1"/>',
                        data,
                        count=1,
                    )
                target.writestr(entry, data)
        return target_bytes.getvalue()

    monkeypatch.setattr(source_domain, "MAX_XLSX_EXTRACT_COLUMNS", 1)
    with pytest.raises(HTTPException, match="source exceeds safe extraction limits"):
        source_domain.extract_blocks("wide.xlsx", workbook_bytes(column=2))

    with pytest.raises(HTTPException, match="source exceeds safe extraction limits"):
        source_domain.extract_blocks(
            "false-dimension.xlsx", declare_only_a1(workbook_bytes(column=2))
        )

    assert "evidence" in {
        block["text"]
        for block in source_domain.extract_blocks(
            "false-row-dimension.xlsx", declare_only_a1(workbook_bytes(row=2))
        )
    }

    monkeypatch.setattr(source_domain, "MAX_XLSX_EXTRACT_WORKSHEETS", 1)
    with pytest.raises(HTTPException, match="source exceeds safe extraction limits"):
        source_domain.extract_blocks("many-sheets.xlsx", workbook_bytes(sheets=2))

    monkeypatch.setattr(source_domain, "MAX_SOURCE_LINE", 8)
    with pytest.raises(HTTPException, match="source exceeds safe extraction limits"):
        source_domain.extract_blocks("long.txt", b"adverse covenant")

    monkeypatch.setattr(source_domain, "MAX_SOURCE_LINE", 20_000)
    monkeypatch.setattr(source_domain, "MAX_SOURCE_TEXT", 13)
    assert [
        block["text"]
        for block in source_domain.extract_blocks("bounded.txt", b"one\ntwo\nthree")
    ] == ["one", "two", "three"]

    monkeypatch.setattr(source_domain, "MAX_SOURCE_TEXT", 8)
    with pytest.raises(HTTPException, match="source exceeds safe extraction limits"):
        source_domain.extract_blocks("long.txt", b"one\ntwo\nthree")


def test_evidence_free_json_values_are_rejected_before_source_set_creation(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    uploads = [
        ("object.json", b"{}"),
        ("list.json", b"[]"),
        ("empty.json", b'""'),
        ("spaces.json", b'" \\t"'),
        ("null.json", b"null"),
    ]

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        for filename, content in uploads:
            case_id = client.post(
                "/api/cases",
                json={"name": filename, "issuer": "Issuer", "sector": "Test"},
            ).json()["id"]
            upload = client.post(
                f"/api/cases/{case_id}/sources",
                files={"file": (filename, content, "application/json")},
            )

            assert upload.status_code == 422
            assert upload.json()["detail"] == "source contains no extractable evidence"
            assert (
                client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
                == "NEEDS_SOURCE"
            )


def test_non_finite_json_values_are_rejected_before_source_set_creation(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    uploads = [
        ("nan.json", b'{"amount":NaN}'),
        ("positive.json", b'{"amount":Infinity}'),
        ("negative.json", b'{"amount":-Infinity}'),
        # Not the NaN/Infinity literals: finite-looking tokens that overflow the double
        # range. `parse_constant` never sees these, so they used to land as `inf`.
        ("overflow.json", b'{"amount":1e999}'),
        ("negative-overflow.json", b'{"amount":-1e999}'),
        ("uppercase-overflow.json", b'{"amount":1E400}'),
    ]

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        for filename, content in uploads:
            case_id = client.post(
                "/api/cases",
                json={"name": filename, "issuer": "Issuer", "sector": "Test"},
            ).json()["id"]
            upload = client.post(
                f"/api/cases/{case_id}/sources",
                files={"file": (filename, content, "application/json")},
            )

            assert upload.status_code == 422
            assert upload.json()["detail"] == "invalid JSON source"
            assert (
                client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
                == "NEEDS_SOURCE"
            )


def test_duplicate_json_keys_are_rejected_before_source_set_creation(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    uploads = [
        ("duplicate.json", b'{"debt":10,"debt":20}'),
        ("nested.json", b'{"facility":{"debt":10,"debt":20}}'),
    ]

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        for filename, content in uploads:
            case_id = client.post(
                "/api/cases",
                json={"name": filename, "issuer": "Issuer", "sector": "Test"},
            ).json()["id"]
            upload = client.post(
                f"/api/cases/{case_id}/sources",
                files={"file": (filename, content, "application/json")},
            )

            assert upload.status_code == 422
            assert upload.json()["detail"] == "invalid JSON source"
            assert (
                client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
                == "NEEDS_SOURCE"
            )


def test_non_utf8_text_sources_are_rejected_before_source_set_creation(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        for filename in ["binary.txt", "binary.md", "binary.csv"]:
            case_id = client.post(
                "/api/cases",
                json={"name": filename, "issuer": "Issuer", "sector": "Test"},
            ).json()["id"]
            upload = client.post(
                f"/api/cases/{case_id}/sources",
                files={"file": (filename, bytes([128]), "text/plain")},
            )

            assert upload.status_code == 422
            assert upload.json()["detail"] == "text source must be UTF-8"
            assert (
                client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
                == "NEEDS_SOURCE"
            )


def test_invalid_pdf_is_rejected_before_source_set_creation(tmp_path: Path) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Invalid PDF", "issuer": "Issuer", "sector": "Test"},
        ).json()["id"]
        upload = client.post(
            f"/api/cases/{case_id}/sources",
            files={"file": ("not-a-pdf.pdf", b"not a PDF", "application/pdf")},
        )

        assert upload.status_code == 422
        assert upload.json()["detail"] == "invalid PDF source"
        assert (
            client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"]
            == "NEEDS_SOURCE"
        )


def test_valid_no_text_pdf_is_accepted(tmp_path: Path) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    source = io.BytesIO()
    writer.write(source)

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Scanned PDF", "issuer": "Issuer", "sector": "Test"},
        ).json()["id"]
        upload = client.post(
            f"/api/cases/{case_id}/sources",
            files={"file": ("scan.pdf", source.getvalue(), "application/pdf")},
        )

        assert upload.status_code == 201
        assert client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"] == "READY"


def test_non_empty_xlsx_is_accepted(tmp_path: Path) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    workbook = Workbook()
    workbook.active["A1"] = "Debt"
    source = io.BytesIO()
    workbook.save(source)

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Workbook source", "issuer": "Issuer", "sector": "Test"},
        ).json()["id"]
        upload = client.post(
            f"/api/cases/{case_id}/sources",
            files={
                "file": (
                    "debt.xlsx",
                    source.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert upload.status_code == 201
        assert client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"] == "READY"


def test_json_scalar_values_are_accepted(tmp_path: Path) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        for filename, content in [
            ("zero.json", b"0"),
            ("false.json", b"false"),
            ("finite.json", b'{"amount":1.25}'),
            ("large-finite.json", b'{"amount":1e308}'),
        ]:
            case_id = client.post(
                "/api/cases",
                json={"name": filename, "issuer": "Issuer", "sector": "Test"},
            ).json()["id"]
            upload = client.post(
                f"/api/cases/{case_id}/sources",
                files={"file": (filename, content, "application/json")},
            )

            assert upload.status_code == 201
            assert (
                client.get(f"/api/cases/{case_id}/pathway-fit").json()["fit"] == "READY"
            )


def test_duplicate_active_source_content_is_rejected_without_creating_a_new_source_set(
    tmp_path: Path,
) -> None:
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )

    with TestClient(create_app(settings, MemoryLedgerSet())) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Duplicate source", "issuer": "Issuer", "sector": "Test"},
        ).json()["id"]
        first = client.post(
            f"/api/cases/{case_id}/sources",
            files={"file": ("evidence.txt", b"Debt 100", "text/plain")},
        )
        duplicate = client.post(
            f"/api/cases/{case_id}/sources",
            files={"file": ("replacement-name.txt", b"Debt 100", "text/plain")},
        )

        assert first.status_code == 201
        assert duplicate.status_code == 409
        assert duplicate.json()["detail"] == "source content already active"
        detail = client.get(f"/api/cases/{case_id}").json()
        assert detail["source_count"] == 1
        assert detail["source_set"]["version"] == 1

        assert (
            client.post(
                f"/api/cases/{case_id}/sources/{first.json()['id']}/withdraw"
            ).status_code
            == 200
        )
        reupload = client.post(
            f"/api/cases/{case_id}/sources",
            files={"file": ("evidence.txt", b"Debt 100", "text/plain")},
        )

        assert reupload.status_code == 201
        detail = client.get(f"/api/cases/{case_id}").json()
        assert detail["source_count"] == 1
        assert detail["source_set"]["version"] == 3


def test_clamav_config_keeps_container_health_and_app_tcp_interfaces() -> None:
    config = (
        (Path(__file__).parents[1] / "deploy" / "clamd.conf").read_text().splitlines()
    )

    assert "LocalSocket /tmp/clamd.sock" in config
    assert "TCPSocket 3310" in config
