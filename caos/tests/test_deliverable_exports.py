from __future__ import annotations

import copy
import io
import hashlib
import os
import threading
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from caos.atomic_files import (
    VaultFileIntegrityError,
    VaultFileUnavailable,
    read_verified_vault_bytes,
)
from caos.contracts import DeliverableDraftRequest, digest
from caos.config import Settings
from caos.http import create_app
from caos.memory_ledgers import MemoryLedgerSet
from caos.postgres_ledgers import PostgresLedgerSet
from caos.publishing.domain import DeliverableService
from caos.publishing import frozen as frozen_publication
from caos.publishing.frozen import DeliverablePublicationService
from caos.publishing.renderers import (
    RENDERER_IDENTITY,
    render_frozen_markdown,
    render_frozen_pdf,
    render_frozen_xlsx,
)
from caos.publishing.templates import DELIVERABLE_TEMPLATES


@pytest.fixture(params=["memory", "postgres"])
def publication_ledgers(request: pytest.FixtureRequest) -> Any:
    if request.param == "memory":
        yield MemoryLedgerSet()
        return
    database_url = os.getenv("CAOS_TEST_DATABASE_URL")
    if not database_url:
        pytest.skip("CAOS_TEST_DATABASE_URL is required for PostgreSQL ledger proof")
    import psycopg

    ledgers = PostgresLedgerSet(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute("TRUNCATE cases RESTART IDENTITY CASCADE")
    try:
        yield ledgers
    finally:
        with psycopg.connect(database_url) as connection:
            with connection.cursor() as cursor:
                cursor.execute("TRUNCATE cases RESTART IDENTITY CASCADE")


def _frozen_payload(pathway: str) -> dict[str, object]:
    template = DELIVERABLE_TEMPLATES[pathway]
    blocks: list[dict[str, object]] = []
    for block in template["blocks"]:
        if block["kind"] == "EVIDENCE_REGISTER":
            blocks.append(
                {
                    "kind": "EVIDENCE_REGISTER",
                    "block_id": block["block_id"],
                    "slot_id": block["slot_id"],
                    "citations": [
                        {
                            "source_id": "source_1",
                            "block_ids": ["b00001"],
                            "claim": "Audited evidence supports this conclusion.",
                        }
                    ],
                }
            )
        else:
            blocks.append(
                {
                    "kind": "NARRATIVE",
                    "block_id": block["block_id"],
                    "slot_id": block["slot_id"],
                    "text": f"{block['title']} includes an exact analyst conclusion.",
                    "content_mode": "ANALYST_JUDGMENT",
                    "citations": [],
                }
            )
    blocks.append(
        {
            "kind": "LIMITATIONS",
            "block_id": f"{pathway.lower()}.limitations",
            "slot_id": "appendix.limitations.01",
            "text": "Model outputs remain sensitive to refinancing assumptions.",
            "citations": [],
        }
    )
    content = {
        "template_id": template["template_id"],
        "template_version": template["template_version"],
        "model_selection": None,
        "model_identity": None,
        "blocks": blocks,
        "generated_blocks": {},
    }
    return {
        "schema_version": "caos.frozen-deliverable.v1",
        "case_id": "case_1",
        "pathway": pathway,
        "draft": {"id": "deliverable_1", "version": 1, "digest": digest(content)},
        "template": {
            "template_id": template["template_id"],
            "template_version": template["template_version"],
            "title": template["title"],
        },
        "authority": {
            "accepted_snapshot_id": "snapshot_1",
            "accepted_snapshot_digest": "a" * 64,
            "source_set_id": "source_set_1",
            "source_set_digest": "b" * 64,
        },
        "model": None,
        "content": content,
        "evidence": [
            {
                "source_id": "source_1",
                "sha256": "c" * 64,
                "block_ids": ["b00001"],
                "withdrawn": False,
            }
        ],
        "methodology": {"build_id": "deploy-v-test"},
        "renderer": RENDERER_IDENTITY,
        "input_fingerprint": "d" * 64,
        "preview_digest": "e" * 64,
    }


def test_all_six_pathways_render_substantive_semantic_exports(tmp_path: Path) -> None:
    for pathway, template in DELIVERABLE_TEMPLATES.items():
        payload = _frozen_payload(pathway)
        markdown = render_frozen_markdown(payload)
        pdf = render_frozen_pdf(payload)
        xlsx = render_frozen_xlsx(payload)

        assert template["title"] in markdown.decode("utf-8")
        assert "refinancing assumptions" in markdown.decode("utf-8")

        reader = PdfReader(io.BytesIO(pdf))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        assert len(reader.pages) >= 2
        assert template["title"] in text
        assert "refinancing assumptions" in text

        workbook_path = tmp_path / f"{pathway}.xlsx"
        workbook_path.write_bytes(xlsx)
        workbook = load_workbook(workbook_path, data_only=False)
        assert {
            "Cover",
            "Reviewed Deliverable",
            "Section Summary",
            "Evidence Register",
            "Revision Record",
        }.issubset(workbook.sheetnames)
        values = {
            cell.value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        }
        assert template["title"] in values
        assert "Model outputs remain sensitive to refinancing assumptions." in values


def test_freeze_pins_and_renders_exact_active_revision_model_authority(
    tmp_path: Path,
) -> None:
    ledgers = MemoryLedgerSet()
    case = ledgers.runs.create_case("Frozen model", "Issuer", "Testing", "analyst")
    authority = _accept_source_authority(ledgers, case["id"], "4")
    outputs = {
        "BASE": {
            "FY2027": {
                "revenue": 825.0,
                "adjusted_ebitda_calc": 150.0,
                "total_debt_reported": 630.0,
                "net_debt": 590.0,
                "total_leverage": 4.2,
                "covenant_headroom": None,
            }
        },
        "BASE_first_breaches": [],
        "DOWNSIDE": {
            "FY2027": {
                "revenue": 760.0,
                "adjusted_ebitda_calc": 125.0,
                "total_debt_reported": 630.0,
                "net_debt": 605.0,
                "total_leverage": 5.04,
                "covenant_headroom": None,
            }
        },
        "DOWNSIDE_first_breaches": [],
    }
    assumptions = [
        {
            "assumption_id": "operating.revenue_growth.consolidated",
            "case": "BASE",
            "period_id": "FY2027",
            "status": "READY",
            "value": 0.03,
            "unit": "PERCENT",
        },
        {
            "assumption_id": "credit.covenant.limit",
            "case": "DOWNSIDE",
            "period_id": "FY2027",
            "status": "UNAVAILABLE",
            "value": None,
            "unit": "MULTIPLE",
            "gap_code": "COVENANT_NOT_DISCLOSED",
        },
    ]

    snapshot = ledgers.runs.get_snapshot(authority["accepted_snapshot_id"])
    assert snapshot is not None
    model_payload = {
        "schema_version": "caos.model.worksheet.v1",
        "identity": {
            "issuer_id": "issuer",
            "issuer_name": "Issuer",
            "analysis_date": "2026-08-26",
        },
        "tabs": [
            {
                "id": "DEBT_SCHEDULE",
                "name": "Debt Schedule",
                "max_row": 1,
                "max_column": 1,
                "freeze_panes": "",
                "merged_cells": [],
                "columns": [
                    {"column": 1, "letter": "A", "width": 12.0, "hidden": False}
                ],
                "cells": [],
            }
        ],
    }
    model_qa = {
        "status": "PASS",
        "semantic_checks": [],
        "semantic_check_count": 0,
        "formula_count": 0,
        "worksheet_cell_count": 0,
        "limitation_flags": ["COVENANT_DATA_UNAVAILABLE"],
        "validation_warnings": ["Covenant headroom cannot be calculated."],
        "source_manifest": [],
    }
    queued, created = ledgers.models.queue_build(
        {
            "case_id": case["id"],
            "accepted_run_id": snapshot["run_id"],
            "accepted_snapshot_id": snapshot["id"],
            "source_set_id": snapshot["source_set_id"],
            "input_fingerprint": "a" * 64,
            "worksheet_schema_version": "caos.model.worksheet.v1",
            "calculation_runtime": {
                "assumption_registry_version": "registry-v1",
                "assumption_registry_digest": "c" * 64,
                "calculation_contract_version": "calculation-v1",
            },
        },
        "analyst",
    )
    assert created is True
    token = ledgers.models.claim(queued["id"], "model-worker")
    assert token is not None
    build = ledgers.models.complete(
        queued["id"],
        token,
        {"payload": model_payload, "payload_digest": digest(model_payload), "qa": model_qa},
        "model-worker",
    )
    revision = ledgers.models.sign_off_revision(
        {
            "case_id": case["id"],
            "build_id": build["id"],
            "accepted_snapshot_id": build["accepted_snapshot_id"],
            "build_input_fingerprint": build["input_fingerprint"],
            "build_payload_digest": build["payload_digest"],
            "registry_version": "registry-v1",
            "registry_digest": "c" * 64,
            "calculation_contract_version": "calculation-v1",
            "effective_assumptions": assumptions,
            "assumptions_digest": digest(assumptions),
            "outputs": outputs,
            "outputs_digest": digest(outputs),
            "preview_digest": "d" * 64,
            "parent_revision_id": None,
            "note": "Quarterly earnings review",
        },
        "analyst",
        expected_head_revision_id=None,
        expected_current_build_id=build["id"],
        expected_current_input_fingerprint=build["input_fingerprint"],
    )
    models = ledgers.models
    authoring = DeliverableService(ledgers.publications, ledgers.sources, models)  # type: ignore[arg-type]
    request = _draft_payload("FULL_CREDIT")
    request["model_selection"] = {
        "kind": "ANALYST_REVISION",
        "build_id": build["id"],
        "revision_id": revision["id"],
    }
    saved = authoring.save(
        case["id"],
        "FULL_CREDIT",
        "analyst",
        DeliverableDraftRequest.model_validate(request),
    )["current"]
    publisher = DeliverablePublicationService(
        ledgers.publications,
        ledgers.runs,
        ledgers.sources,
        models,
        authoring,
        SimpleNamespace(build_id="deploy-v-test"),
        tmp_path / "vault",
    )
    frozen = publisher.freeze(
        case["id"],
        "FULL_CREDIT",
        "analyst",
        saved["id"],
        saved["version"],
        saved["digest"],
    )

    pinned = frozen["payload"]["model"]
    assert pinned["outputs"] == outputs
    assert pinned["effective_assumptions"] == assumptions
    assert pinned["application_build"]["payload"] == build["payload"]
    assert pinned["application_build"]["qa"] == build["qa"]

    pdf = (tmp_path / "vault" / frozen["exports"]["pdf"]["vault_key"]).read_bytes()
    pdf_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages
    )
    assert "Base / Downside Model Analysis" in pdf_text
    assert "BASE" in pdf_text
    assert "FY2027" in pdf_text
    assert "total_leverage" in pdf_text
    assert "4.2" in pdf_text

    xlsx = (tmp_path / "vault" / frozen["exports"]["xlsx"]["vault_key"]).read_bytes()
    workbook = load_workbook(io.BytesIO(xlsx), data_only=False)
    assert {
        "Base Downside",
        "Model",
        "Assumptions",
        "Debt Schedule",
        "Gaps and Warnings",
    }.issubset(workbook.sheetnames)
    values = {
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    }
    assert 4.2 in values
    assert "Covenant headroom cannot be calculated." in values


def _frozen_ledger_value(
    draft: dict[str, Any], label: str, authority: dict[str, Any]
) -> dict[str, Any]:
    payload = _frozen_payload("RELATIVE_VALUE")
    payload["case_id"] = draft["case_id"]
    payload["draft"] = {
        "id": draft["id"],
        "version": draft["version"],
        "digest": draft["digest"],
    }
    payload["authority"] = authority
    payload["evidence"] = []
    payload["input_fingerprint"] = digest({"draft": draft["digest"], "label": label})
    payload["preview_digest"] = digest(
        {key: value for key, value in payload.items() if key != "preview_digest"}
    )
    return {
        "payload": payload,
        "digest": digest(payload),
        "preview_digest": payload["preview_digest"],
        "input_fingerprint": payload["input_fingerprint"],
        "authority_identity": payload["authority"],
        "model_identity": payload["model"],
        "template_identity": payload["template"],
        "render_identity": payload["renderer"],
    }


def _export_metadata(label: str) -> dict[str, dict[str, Any]]:
    return {
        format_name: {
            "format": format_name,
            "vault_key": f"deliverables/case/{label}/{label}.{format_name}",
            "sha256": format_name[0] * 64,
            "size": 100 + index,
            "renderer_identity": RENDERER_IDENTITY,
        }
        for index, format_name in enumerate(("md", "pdf", "xlsx"), start=1)
    }


def _draft_payload(pathway: str = "RELATIVE_VALUE") -> dict[str, Any]:
    template = DELIVERABLE_TEMPLATES[pathway]
    return {
        "expected_version": 0,
        "template_id": template["template_id"],
        "template_version": template["template_version"],
        "model_selection": None,
        "blocks": [
            {
                "kind": "NARRATIVE",
                "block_id": block["block_id"],
                "slot_id": block["slot_id"],
                "text": f"{block['title']} committee conclusion.",
                "content_mode": "ANALYST_JUDGMENT",
                "citations": [],
            }
            if block["kind"] == "NARRATIVE"
            else {
                "kind": "EVIDENCE_REGISTER",
                "block_id": block["block_id"],
                "slot_id": block["slot_id"],
                "citations": [],
            }
            for block in template["blocks"]
        ],
    }


def _accept_source_authority(
    ledgers: Any, case_id: str, suffix: str = "9"
) -> dict[str, Any]:
    ingested = ledgers.sources.ingest(
        {
            "case_id": case_id,
            "filename": f"earnings-{suffix}.txt",
            "media_type": "text/plain",
            "bytes": 8,
            "sha256": suffix * 64,
            "vault_path": f"/tmp/non-reading-source-{suffix}",
            "blocks": [
                {
                    "block_id": "b00001",
                    "locator": {"line": 1},
                    "text": "Evidence",
                    "extractor_version": "test-v1",
                    "confidence": "HIGH",
                    "untrusted_data": True,
                }
            ],
            "created_by": "local-analyst",
            "created_at": "2026-08-26T00:00:00+00:00",
            "withdrawn": False,
        },
        "local-analyst",
    )
    source_set = ingested["source_set"]
    run = ledgers.runs.create_run_with_nodes(
        case_id,
        "local-analyst",
        {"pathway": "EARNINGS_UPDATE", "source_set_id": source_set["id"]},
        [],
    )
    token = ledgers.runs.claim(run["id"], "worker")
    assert token is not None
    ledgers.runs.finalize_success(run["id"], token, None, {"run_id": run["id"]})
    snapshot = {
        "case_id": case_id,
        "run_id": run["id"],
        "source_set_id": source_set["id"],
        "source_set_version": source_set["version"],
        "artifacts": [],
        "accepted_at": "2026-08-26T00:00:00+00:00",
    }
    accepted = ledgers.runs.accept_snapshot(
        case_id,
        run["id"],
        "local-analyst",
        {**snapshot, "digest": digest(snapshot)},
    )
    return {
        "accepted_snapshot_id": accepted["id"],
        "accepted_snapshot_digest": digest(accepted),
        "source_set_id": source_set["id"],
        "source_set_version": source_set["version"],
        "source_set_digest": digest(source_set),
    }


def test_frozen_filed_history_and_request_changes_are_atomic(
    publication_ledgers: Any,
) -> None:
    ledgers = publication_ledgers
    case = ledgers.runs.create_case("Deliverable", "Issuer", "Testing", "analyst")
    authority = _accept_source_authority(ledgers, case["id"])
    draft = ledgers.publications.append_deliverable_revision(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        0,
        {
            "template_id": "caos.relative-value.v1",
            "template_version": "caos.deliverable-template.v1",
            "digest": "a" * 64,
            "content": _frozen_payload("RELATIVE_VALUE")["content"],
        },
    )
    frozen_value = _frozen_ledger_value(draft, "first", authority)
    frozen = ledgers.publications.append_frozen_deliverable(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        draft["id"],
        draft["version"],
        draft["digest"],
        frozen_value,
        _export_metadata("first"),
    )
    assert frozen["status"] == "FROZEN"
    assert set(frozen["exports"]) == {"md", "pdf", "xlsx"}

    def file_once(actor: str) -> tuple[str, Any]:
        try:
            return (
                "filed",
                ledgers.publications.file_deliverable(
                    case["id"],
                    frozen["id"],
                    actor,
                    frozen["preview_digest"],
                    frozen["input_fingerprint"],
                ),
            )
        except ValueError as exc:
            return "conflict", exc

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(file_once, ("approver-a", "approver-b")))
    assert [status for status, _ in results].count("filed") == 1
    assert [status for status, _ in results].count("conflict") == 1
    filed = next(value for status, value in results if status == "filed")
    assert filed["status"] == "FILED"
    assert ledgers.publications.get_frozen_deliverable(filed["id"])["payload"] == frozen["payload"]

    second_draft = ledgers.publications.append_deliverable_revision(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        draft["version"],
        {
            "template_id": draft["template_id"],
            "template_version": draft["template_version"],
            "digest": "b" * 64,
            "content": draft["content"],
        },
    )
    second_value = _frozen_ledger_value(second_draft, "second", authority)
    second_frozen = ledgers.publications.append_frozen_deliverable(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        second_draft["id"],
        second_draft["version"],
        second_draft["digest"],
        second_value,
        _export_metadata("second"),
    )
    changed, replacement = ledgers.publications.request_deliverable_changes(
        case["id"],
        second_frozen["id"],
        "approver-a",
        second_frozen["preview_digest"],
        second_frozen["input_fingerprint"],
        "Clarify the downside liquidity conclusion.",
    )
    assert changed["status"] == "CHANGES_REQUESTED"
    assert replacement["version"] == second_draft["version"] + 1
    assert replacement["content"] == second_draft["content"]
    assert replacement["change_request"]["comment"] == "Clarify the downside liquidity conclusion."

    third_value = _frozen_ledger_value(replacement, "third", authority)
    third_frozen = ledgers.publications.append_frozen_deliverable(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        replacement["id"],
        replacement["version"],
        replacement["digest"],
        third_value,
        _export_metadata("third"),
    )
    third_filed = ledgers.publications.file_deliverable(
        case["id"],
        third_frozen["id"],
        "approver-a",
        third_frozen["preview_digest"],
        third_frozen["input_fingerprint"],
    )
    assert third_filed["status"] == "FILED"
    prior = ledgers.publications.get_frozen_deliverable(filed["id"])
    assert prior["status"] == "SUPERSEDED"
    assert prior["superseded_by_id"] == third_frozen["id"]
    assert set(prior["exports"]) == {"md", "pdf", "xlsx"}

    actions = [
        event["action"]
        for event in ledgers.publications.list_audit()
        if event.get("case_id") == case["id"]
    ]
    assert actions.count("deliverable.frozen") == 3
    assert actions.count("deliverable.filed") == 2
    assert actions.count("deliverable.changes_requested") == 1


def test_exact_draft_freeze_is_one_idempotent_record_under_race(
    publication_ledgers: Any,
) -> None:
    ledgers = publication_ledgers
    case = ledgers.runs.create_case("Freeze race", "Issuer", "Testing", "analyst")
    authority = _accept_source_authority(ledgers, case["id"], "5")
    draft = ledgers.publications.append_deliverable_revision(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        0,
        {
            "template_id": "caos.relative-value.v1",
            "template_version": "caos.deliverable-template.v1",
            "digest": "a" * 64,
            "content": _frozen_payload("RELATIVE_VALUE")["content"],
        },
    )
    frozen_value = _frozen_ledger_value(draft, "same-draft", authority)
    exports = _export_metadata("same-draft")
    barrier = threading.Barrier(2)

    def freeze_once(_index: int) -> dict[str, Any]:
        barrier.wait(timeout=5)
        return ledgers.publications.append_frozen_deliverable(
            case["id"],
            "RELATIVE_VALUE",
            "analyst",
            draft["id"],
            draft["version"],
            draft["digest"],
            frozen_value,
            exports,
        )

    with ThreadPoolExecutor(max_workers=2) as pool:
        raced = list(pool.map(freeze_once, range(2)))
    assert len({record["id"] for record in raced}) == 1
    retry = ledgers.publications.append_frozen_deliverable(
        case["id"],
        "RELATIVE_VALUE",
        "analyst",
        draft["id"],
        draft["version"],
        draft["digest"],
        frozen_value,
        exports,
    )
    assert retry["id"] == raced[0]["id"]

    history = ledgers.publications.list_frozen_deliverables(
        case["id"], "RELATIVE_VALUE"
    )
    assert [record["id"] for record in history] == [raced[0]["id"]]
    assert set(history[0]["exports"]) == {"md", "pdf", "xlsx"}
    actions = [
        event["action"]
        for event in ledgers.publications.list_audit()
        if event.get("case_id") == case["id"]
    ]
    assert actions.count("deliverable.frozen") == 1

    inconsistent_exports = copy.deepcopy(exports)
    inconsistent_exports["pdf"]["sha256"] = "0" * 64
    with pytest.raises(ValueError, match="DELIVERABLE_FREEZE_CONFLICT"):
        ledgers.publications.append_frozen_deliverable(
            case["id"],
            "RELATIVE_VALUE",
            "analyst",
            draft["id"],
            draft["version"],
            draft["digest"],
            frozen_value,
            inconsistent_exports,
        )
    filed = ledgers.publications.file_deliverable(
        case["id"],
        retry["id"],
        "approver",
        retry["preview_digest"],
        retry["input_fingerprint"],
    )
    assert filed["status"] == "FILED"
    assert [
        record["status"]
        for record in ledgers.publications.list_frozen_deliverables(
            case["id"], "RELATIVE_VALUE"
        )
    ] == ["FILED"]


def test_http_freeze_file_and_download_exact_stored_bytes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ledgers = MemoryLedgerSet()
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    app = create_app(settings, ledgers)
    with TestClient(app) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Report", "issuer": "Issuer", "sector": "Testing"},
        ).json()["id"]
        _accept_source_authority(ledgers, case_id)
        route = f"/api/cases/{case_id}/deliverables/RELATIVE_VALUE"
        current = client.put(f"{route}/draft", json=_draft_payload()).json()["current"]
        frozen_response = client.post(
            f"{route}/freeze",
            json={
                "draft_id": current["id"],
                "draft_version": current["version"],
                "draft_digest": current["digest"],
            },
        )
        assert frozen_response.status_code == 201, frozen_response.text
        frozen = frozen_response.json()
        duplicate = client.post(
            f"{route}/freeze",
            json={
                "draft_id": current["id"],
                "draft_version": current["version"],
                "draft_digest": current["digest"],
            },
        )
        assert duplicate.status_code == 201
        assert duplicate.json()["id"] == frozen["id"]

        original_renderer = frozen_publication.render_frozen_exports

        def inconsistent_renderer(payload: dict[str, Any]) -> dict[str, bytes]:
            rendered = original_renderer(payload)
            return {**rendered, "md": rendered["md"] + b"\n"}

        with monkeypatch.context() as conflict_patch:
            conflict_patch.setattr(
                "caos.publishing.frozen.render_frozen_exports",
                inconsistent_renderer,
            )
            conflict = client.post(
                f"{route}/freeze",
                json={
                    "draft_id": current["id"],
                    "draft_version": current["version"],
                    "draft_digest": current["digest"],
                },
            )
        assert conflict.status_code == 409
        assert conflict.json()["detail"]["code"] == "DELIVERABLE_FREEZE_CONFLICT"
        assert conflict.json()["detail"]["current"]["id"] == frozen["id"]
        assert [
            event["action"]
            for event in ledgers.publications.list_audit()
            if event.get("case_id") == case_id
        ].count("deliverable.frozen") == 1
        export_route = (
            f"/api/cases/{case_id}/deliverables/by-id/{frozen['id']}/export"
        )
        assert client.get(f"{export_route}/pdf").status_code == 409
        assert client.post(
            f"/api/cases/{case_id}/deliverables/by-id/{frozen['id']}/approve",
            headers={"x-caos-role": "ANALYST"},
            json={
                "preview_digest": frozen["preview_digest"],
                "input_fingerprint": frozen["input_fingerprint"],
            },
        ).status_code == 403
        filed_response = client.post(
            f"/api/cases/{case_id}/deliverables/by-id/{frozen['id']}/approve",
            headers={"x-caos-role": "APPROVER"},
            json={
                "preview_digest": frozen["preview_digest"],
                "input_fingerprint": frozen["input_fingerprint"],
            },
        )
        assert filed_response.status_code == 200, filed_response.text
        filed = filed_response.json()
        assert filed["status"] == "FILED"
        for format_name in ("md", "pdf", "xlsx"):
            downloaded = client.get(f"{export_route}/{format_name}")
            assert downloaded.status_code == 200
            assert hashlib.sha256(downloaded.content).hexdigest() == filed["exports"][
                format_name
            ]["sha256"]
            assert len(downloaded.content) == filed["exports"][format_name]["size"]

        pdf_path = tmp_path / "vault" / filed["exports"]["pdf"]["vault_key"]
        original = pdf_path.read_bytes()
        assert client.get(f"{export_route}/pdf").content == original

        def renderer_must_not_run(_payload: dict[str, Any]) -> dict[str, bytes]:
            raise AssertionError("download must not rerender a filed deliverable")

        monkeypatch.setattr(
            "caos.publishing.frozen.render_frozen_exports", renderer_must_not_run
        )
        with TestClient(create_app(settings, ledgers)) as restarted:
            assert restarted.get(f"{export_route}/pdf").content == original

        final_open_reached = threading.Event()
        allow_final_open = threading.Event()
        original_open = os.open

        def paused_final_open(
            path: str | bytes | os.PathLike[str] | os.PathLike[bytes],
            flags: int,
            mode: int = 0o777,
            *,
            dir_fd: int | None = None,
        ) -> int:
            if path == pdf_path.name and dir_fd is not None:
                final_open_reached.set()
                assert allow_final_open.wait(timeout=5)
            return original_open(path, flags, mode, dir_fd=dir_fd)

        audit_count = len(ledgers.publications.list_audit())
        with monkeypatch.context() as race_patch:
            race_patch.setattr("caos.atomic_files.os.open", paused_final_open)
            with ThreadPoolExecutor(max_workers=1) as pool:
                request = pool.submit(client.get, f"{export_route}/pdf")
                assert final_open_reached.wait(timeout=2)
                outside_pdf = tmp_path / "outside-identical.pdf"
                pdf_path.replace(outside_pdf)
                pdf_path.symlink_to(outside_pdf)
                allow_final_open.set()
                raced = request.result(timeout=5)
        assert raced.status_code == 409
        assert raced.json()["detail"] == "DELIVERABLE_EXPORT_UNAVAILABLE"
        assert len(ledgers.publications.list_audit()) == audit_count
        pdf_path.unlink()
        outside_pdf.replace(pdf_path)

        pdf_path.write_bytes(original + b"tamper")
        assert client.get(f"{export_route}/pdf").json()["detail"] == (
            "DELIVERABLE_EXPORT_INTEGRITY_FAILED"
        )


def test_secure_vault_reader_rejects_component_swaps_oversize_and_nonregular(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root = tmp_path / "vault"
    directory = root / "deliverables" / "case"
    directory.mkdir(parents=True)
    target = directory / "export.pdf"
    content = b"exact frozen bytes"
    target.write_bytes(content)
    checksum = hashlib.sha256(content).hexdigest()
    key = "deliverables/case/export.pdf"

    def read() -> bytes:
        return read_verified_vault_bytes(
            root,
            key,
            expected_sha256=checksum,
            expected_size=len(content),
            max_bytes=1024,
        )

    assert read() == content
    with monkeypatch.context() as unsupported_patch:
        unsupported_patch.delattr(os, "O_NOFOLLOW")
        with pytest.raises(VaultFileUnavailable, match="no-follow"):
            read()
    target.write_bytes(content + b"oversized")
    with pytest.raises(VaultFileIntegrityError, match="size mismatch"):
        read()

    target.unlink()
    target.symlink_to("/dev/null")
    with pytest.raises(VaultFileUnavailable):
        read()
    target.unlink()
    os.mkfifo(target)
    with pytest.raises(VaultFileUnavailable, match="regular file"):
        read()
    target.unlink()
    target.write_bytes(content)

    component_open_reached = threading.Event()
    allow_component_open = threading.Event()
    original_open = os.open

    def paused_component_open(
        path: str | bytes | os.PathLike[str] | os.PathLike[bytes],
        flags: int,
        mode: int = 0o777,
        *,
        dir_fd: int | None = None,
    ) -> int:
        if path == "case" and dir_fd is not None:
            component_open_reached.set()
            assert allow_component_open.wait(timeout=5)
        return original_open(path, flags, mode, dir_fd=dir_fd)

    outside = tmp_path / "outside"
    outside.mkdir()
    (outside / target.name).write_bytes(content)
    moved = root / "deliverables" / "moved-case"
    with monkeypatch.context() as race_patch:
        race_patch.setattr("caos.atomic_files.os.open", paused_component_open)
        with ThreadPoolExecutor(max_workers=1) as pool:
            request = pool.submit(read)
            assert component_open_reached.wait(timeout=2)
            directory.replace(moved)
            directory.symlink_to(outside, target_is_directory=True)
            allow_component_open.set()
            with pytest.raises(VaultFileUnavailable):
                request.result(timeout=5)


def test_approval_revalidates_current_accepted_authority_without_residue(
    tmp_path: Path,
) -> None:
    ledgers = MemoryLedgerSet()
    app = create_app(
        Settings(
            storage_dir=tmp_path / "vault",
            deploy_v_root=Path(__file__).parents[1]
            / "server"
            / "caos"
            / "methodology"
            / "vendor"
            / "deploy_v",
        ),
        ledgers,
    )
    with TestClient(app) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Stale", "issuer": "Issuer", "sector": "Testing"},
        ).json()["id"]
        _accept_source_authority(ledgers, case_id, "7")
        route = f"/api/cases/{case_id}/deliverables/RELATIVE_VALUE"
        current = client.put(f"{route}/draft", json=_draft_payload()).json()["current"]
        frozen = client.post(
            f"{route}/freeze",
            json={
                "draft_id": current["id"],
                "draft_version": current["version"],
                "draft_digest": current["digest"],
            },
        ).json()

        _accept_source_authority(ledgers, case_id, "8")
        rejected = client.post(
            f"/api/cases/{case_id}/deliverables/by-id/{frozen['id']}/approve",
            headers={"x-caos-role": "APPROVER"},
            json={
                "preview_digest": frozen["preview_digest"],
                "input_fingerprint": frozen["input_fingerprint"],
            },
        )
        assert rejected.status_code == 409
        assert rejected.json()["detail"] == "DELIVERABLE_FROZEN_AUTHORITY_STALE"
        stored = ledgers.publications.get_frozen_deliverable(frozen["id"])
        assert stored["status"] == "FROZEN"
        assert not [
            event
            for event in ledgers.publications.list_audit()
            if event["action"] == "deliverable.filed"
            and event.get("case_id") == case_id
        ]


def test_http_request_changes_requires_approver_comment_and_appends_draft(
    tmp_path: Path,
) -> None:
    ledgers = MemoryLedgerSet()
    settings = Settings(
        storage_dir=tmp_path / "vault",
        deploy_v_root=Path(__file__).parents[1]
        / "server"
        / "caos"
        / "methodology"
        / "vendor"
        / "deploy_v",
    )
    with TestClient(create_app(settings, ledgers)) as client:
        case_id = client.post(
            "/api/cases",
            json={"name": "Changes", "issuer": "Issuer", "sector": "Testing"},
        ).json()["id"]
        _accept_source_authority(ledgers, case_id, "6")
        route = f"/api/cases/{case_id}/deliverables/RELATIVE_VALUE"
        draft = client.put(f"{route}/draft", json=_draft_payload()).json()["current"]
        frozen = client.post(
            f"{route}/freeze",
            json={
                "draft_id": draft["id"],
                "draft_version": draft["version"],
                "draft_digest": draft["digest"],
            },
        ).json()
        change_route = (
            f"/api/cases/{case_id}/deliverables/by-id/"
            f"{frozen['id']}/request-changes"
        )
        identity = {
            "preview_digest": frozen["preview_digest"],
            "input_fingerprint": frozen["input_fingerprint"],
        }
        assert client.post(
            change_route,
            headers={"x-caos-role": "APPROVER"},
            json={**identity, "comment": "   "},
        ).status_code == 422
        assert client.post(
            change_route,
            headers={"x-caos-role": "ANALYST"},
            json={**identity, "comment": "Clarify downside."},
        ).status_code == 403
        changed = client.post(
            change_route,
            headers={"x-caos-role": "APPROVER"},
            json={**identity, "comment": "Clarify downside."},
        )
        assert changed.status_code == 200, changed.text
        assert changed.json()["frozen"]["status"] == "CHANGES_REQUESTED"
        assert changed.json()["draft"]["version"] == 2
        assert changed.json()["draft"]["change_request"]["comment"] == (
            "Clarify downside."
        )


def test_xlsx_neutralizes_formula_text_and_preserves_typed_model_values(
    tmp_path: Path,
) -> None:
    payload = _frozen_payload("FULL_CREDIT")
    payload["content"]["blocks"][-1]["text"] = '=HYPERLINK("https://invalid", "click")'
    payload["content"]["blocks"].append(
        {
            "kind": "GENERATED_METRIC",
            "block_id": "full-credit.metric",
            "slot_id": "appendix.generated-metric.01",
            "metric_ids": ["total_leverage"],
        }
    )
    payload["content"]["generated_blocks"] = {
        "full-credit.metric": {
            "kind": "GENERATED_METRIC",
            "outputs": {"BASE": {"FY2027": {"total_leverage": 4.2}}},
        }
    }
    rendered = render_frozen_xlsx(payload)
    workbook_path = tmp_path / "safe.xlsx"
    workbook_path.write_bytes(rendered)
    workbook = load_workbook(workbook_path, data_only=False)
    assert "Base Downside" in workbook.sheetnames
    values = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ]
    assert 4.2 in values
    assert '\'=HYPERLINK("https://invalid", "click")' in values
    assert all(cell.data_type != "f" for sheet in workbook for row in sheet for cell in row)
