"""Pure renderers for one immutable canonical Frozen Deliverable payload."""

from __future__ import annotations

import html
import io
import re
import zipfile
from collections.abc import Iterable
from copy import deepcopy
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..contracts import digest


FROZEN_SCHEMA_VERSION = "caos.frozen-deliverable.v1"
RENDERER_VERSION = "caos.deliverable-renderer.v2"
RENDERER_IDENTITY = {
    "version": RENDERER_VERSION,
    "contract_digest": digest(
        {
            "schema": FROZEN_SCHEMA_VERSION,
            "formats": ["md", "pdf", "xlsx"],
            "xlsx_formula_policy": "untrusted-text-only",
        }
    ),
}
# openpyxl always stamps workbook.properties.modified with datetime.now() at save
# time (openpyxl/writer/excel.py:save_workbook, unconditional, not overridable via
# the property) and every ZIP entry's date_time with the wall clock too -- both make
# "frozen" XLSX exports non-byte-reproducible across time (render_frozen_pdf already
# avoids the equivalent reportlab issue via invariant=1). _freeze_zip_timestamps
# rewrites both after the fact. 1980-01-01 is the floor date ZIP timestamps can
# represent, so it's a safe fixed value rather than an arbitrary one.
_DETERMINISTIC_TIMESTAMP = datetime(1980, 1, 1, tzinfo=timezone.utc)
_DETERMINISTIC_ZIP_DATE_TIME = (1980, 1, 1, 0, 0, 0)
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
BASE_SHEETS = (
    "Cover",
    "Reviewed Deliverable",
    "Section Summary",
    "Evidence Register",
    "Revision Record",
)


def _validate_payload(payload: dict[str, Any]) -> None:
    if payload.get("schema_version") != FROZEN_SCHEMA_VERSION:
        raise ValueError("FROZEN_DELIVERABLE_SCHEMA_INVALID")
    if payload.get("renderer") != RENDERER_IDENTITY:
        raise ValueError("FROZEN_DELIVERABLE_RENDERER_INVALID")
    content = payload.get("content")
    if not isinstance(content, dict) or not isinstance(content.get("blocks"), list):
        raise ValueError("FROZEN_DELIVERABLE_CONTENT_INVALID")
    for required in (
        "case_id",
        "pathway",
        "draft",
        "template",
        "authority",
        "methodology",
        "input_fingerprint",
        "preview_digest",
    ):
        if required not in payload:
            raise ValueError("FROZEN_DELIVERABLE_CONTENT_INVALID")


def _title_for(payload: dict[str, Any]) -> str:
    title = payload["template"].get("title")
    return title if isinstance(title, str) and title else payload["pathway"]


def _block_title(block: dict[str, Any]) -> str:
    title = block.get("title")
    if isinstance(title, str) and title:
        return title
    block_id = str(block.get("block_id", "Section"))
    return block_id.rsplit(".", 1)[-1].replace("-", " ").replace("_", " ").title()


def _markdown_text(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def _flatten(value: Any, prefix: str = "") -> Iterable[tuple[str, Any]]:
    if isinstance(value, dict):
        for key in sorted(value):
            nested = f"{prefix}.{key}" if prefix else str(key)
            yield from _flatten(value[key], nested)
    elif isinstance(value, list):
        for index, item in enumerate(value, start=1):
            nested = f"{prefix}[{index}]"
            yield from _flatten(item, nested)
    else:
        yield prefix, value


def _citation_lines(block: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for citation in block.get("citations") or []:
        blocks = ", ".join(citation.get("block_ids") or [])
        lines.append(
            f"{citation.get('claim', 'Evidence')} — source "
            f"{citation.get('source_id')} [{blocks}]"
        )
    return lines


def _model_output_rows(payload: dict[str, Any]) -> list[list[Any]]:
    outputs = (payload.get("model") or {}).get("outputs") or {}
    rows: list[list[Any]] = []
    for case in ("BASE", "DOWNSIDE"):
        periods = outputs.get(case)
        if not isinstance(periods, dict):
            continue
        for period_id in sorted(periods):
            values = periods[period_id]
            if not isinstance(values, dict):
                continue
            rows.extend(
                [case, period_id, output_id, values[output_id]]
                for output_id in sorted(values)
            )
    return rows


def _model_assumption_rows(payload: dict[str, Any]) -> list[list[Any]]:
    assumptions = (payload.get("model") or {}).get("effective_assumptions") or []
    rows = [
        [
            row.get("assumption_id"),
            row.get("case"),
            row.get("period_id"),
            row.get("unit"),
            row.get("status"),
            row.get("value"),
            row.get("gap_code"),
        ]
        for row in assumptions
        if isinstance(row, dict)
    ]
    return sorted(rows, key=lambda row: tuple(str(value or "") for value in row[:3]))


def _model_debt_rows(payload: dict[str, Any]) -> list[list[Any]]:
    debt = (payload.get("model") or {}).get("debt") or {}
    rows: list[list[Any]] = []
    for case in ("BASE", "DOWNSIDE"):
        periods = debt.get(case)
        if not isinstance(periods, dict):
            continue
        for period_id in sorted(periods):
            values = periods[period_id]
            if not isinstance(values, dict):
                continue
            rows.extend(
                [case, period_id, output_id, values[output_id]]
                for output_id in sorted(values)
            )
    return rows


def _model_gap_warning_rows(payload: dict[str, Any]) -> list[list[Any]]:
    model = payload.get("model") or {}
    rows: list[list[Any]] = []
    for gap in model.get("assumption_gaps") or []:
        if isinstance(gap, dict):
            rows.append(
                [
                    "ASSUMPTION_GAP",
                    ".".join(
                        str(gap.get(key) or "")
                        for key in ("assumption_id", "case", "period_id")
                    ),
                    gap.get("gap_code") or gap.get("status"),
                ]
            )
    warnings = model.get("warnings") or {}
    if isinstance(warnings, dict):
        for key in ("limitation_flags", "validation_warnings"):
            rows.extend([key.upper(), key, value] for value in warnings.get(key) or [])
    rows.extend(
        ["MODEL_OUTPUT", f"{case}.{period_id}.{output_id}", "UNAVAILABLE"]
        for case, period_id, output_id, value in _model_output_rows(payload)
        if value is None
    )
    return rows


def render_frozen_markdown(payload: dict[str, Any]) -> bytes:
    """Render human-readable Markdown exclusively from the frozen payload."""

    _validate_payload(payload)
    authority = payload["authority"]
    model = payload.get("model")
    lines = [
        f"# {_title_for(payload)}",
        "",
        f"Pathway: `{payload['pathway']}`",
        f"Frozen preview: `{payload['preview_digest']}`",
        f"Input fingerprint: `{payload['input_fingerprint']}`",
        f"Accepted snapshot: `{authority.get('accepted_snapshot_id')}`",
        f"Source set: `{authority.get('source_set_id')}`",
        f"Model authority: `{model.get('kind')}` / `{model.get('build_id')}`"
        if isinstance(model, dict)
        else "Model authority: not included",
    ]
    model_outputs = _model_output_rows(payload)
    if model_outputs:
        lines.extend(["", "## Base / Downside Model Analysis", ""])
        lines.extend(
            f"- **{case} / {period_id} / {output_id}:** {_markdown_text(value)}"
            for case, period_id, output_id, value in model_outputs
        )
    assumptions = _model_assumption_rows(payload)
    if assumptions:
        lines.extend(["", "## Effective Assumptions", ""])
        lines.extend(
            f"- **{assumption_id} / {case} / {period_id}:** "
            f"{_markdown_text(value)} {unit} ({status}; {gap_code or 'no gap'})"
            for assumption_id, case, period_id, unit, status, value, gap_code in assumptions
        )
    gap_warnings = _model_gap_warning_rows(payload)
    if gap_warnings:
        lines.extend(["", "## Model Gaps and Warnings", ""])
        lines.extend(
            f"- **{_markdown_text(kind)} / {_markdown_text(field)}:** "
            f"{_markdown_text(value)}"
            for kind, field, value in gap_warnings
        )
    generated = payload["content"].get("generated_blocks") or {}
    for block in payload["content"]["blocks"]:
        lines.extend(["", f"## {_block_title(block)}", ""])
        kind = block["kind"]
        if kind in {"NARRATIVE", "LIMITATIONS", "HEADING"}:
            lines.append(str(block.get("text", "")))
        elif kind == "EVIDENCE_REGISTER":
            lines.extend(f"- {_markdown_text(line)}" for line in _citation_lines(block))
        else:
            server_value = generated.get(block["block_id"], {})
            for key, value in _flatten(server_value):
                lines.append(f"- **{_markdown_text(key)}:** {_markdown_text(value)}")
        citations = _citation_lines(block)
        if citations and kind != "EVIDENCE_REGISTER":
            lines.extend(["", "Evidence:"])
            lines.extend(f"- {_markdown_text(line)}" for line in citations)
    lines.extend(
        [
            "",
            "## Revision Record",
            "",
            f"- Draft: `{payload['draft']['id']}` v{payload['draft']['version']}",
            f"- Template: `{payload['template']['template_id']}` "
            f"/ `{payload['template']['template_version']}`",
            f"- Renderer: `{payload['renderer']['version']}`",
            f"- Methodology: `{payload['methodology']['build_id']}`",
        ]
    )
    return ("\n".join(lines) + "\n").encode("utf-8")


class _InvariantCanvas(Canvas):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        kwargs["invariant"] = 1
        super().__init__(*args, **kwargs)


def render_frozen_pdf(payload: dict[str, Any]) -> bytes:
    """Render a deterministic, multi-page committee-ready PDF."""

    _validate_payload(payload)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=LETTER,
        leftMargin=0.72 * inch,
        rightMargin=0.72 * inch,
        topMargin=0.68 * inch,
        bottomMargin=0.68 * inch,
        title=_title_for(payload),
        author="CAOS",
        subject=f"Filed {payload['pathway']} deliverable",
    )
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            "CAOSLabel",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=9,
            textColor=colors.HexColor("#4a4a55"),
            spaceAfter=5,
            alignment=TA_LEFT,
        )
    )
    styles["Title"].fontName = "Helvetica-Bold"
    styles["Title"].fontSize = 20
    styles["Heading2"].fontName = "Helvetica-Bold"
    styles["Heading2"].fontSize = 13
    story: list[Any] = [
        Paragraph(html.escape(_title_for(payload)), styles["Title"]),
        Spacer(1, 0.15 * inch),
        Paragraph(
            html.escape(
                f"{payload['pathway']}  |  FROZEN PREVIEW {payload['preview_digest']}"
            ),
            styles["CAOSLabel"],
        ),
        Table(
            [
                ["Accepted snapshot", payload["authority"].get("accepted_snapshot_id")],
                ["Source set", payload["authority"].get("source_set_id")],
                ["Model build", (payload.get("model") or {}).get("build_id", "Not included")],
                ["Template", payload["template"].get("template_version")],
                ["Methodology", payload["methodology"].get("build_id")],
                ["Input fingerprint", payload["input_fingerprint"]],
            ],
            colWidths=[1.55 * inch, 5.1 * inch],
        ),
        PageBreak(),
    ]
    model_outputs = _model_output_rows(payload)
    if model_outputs:
        story.extend(
            [
                Paragraph("Base / Downside Model Analysis", styles["Heading2"]),
                Table(
                    [["Case", "Period", "Output", "Frozen value"]] + model_outputs,
                    repeatRows=1,
                    colWidths=[0.75 * inch, 0.8 * inch, 2.75 * inch, 2.25 * inch],
                ),
                Spacer(1, 0.18 * inch),
            ]
        )
    assumptions = _model_assumption_rows(payload)
    if assumptions:
        story.extend(
            [
                Paragraph("Effective Assumptions", styles["Heading2"]),
                Table(
                    [["Assumption", "Case", "Period", "Unit", "Status", "Value", "Gap"]]
                    + assumptions,
                    repeatRows=1,
                    colWidths=[1.7 * inch, 0.55 * inch, 0.6 * inch, 0.65 * inch, 0.65 * inch, 0.65 * inch, 1.15 * inch],
                ),
                Spacer(1, 0.18 * inch),
            ]
        )
    gap_warnings = _model_gap_warning_rows(payload)
    if gap_warnings:
        story.extend(
            [
                Paragraph("Model Gaps and Warnings", styles["Heading2"]),
                Table(
                    [["Type", "Field", "Frozen value"]] + gap_warnings,
                    repeatRows=1,
                    colWidths=[1.35 * inch, 2.8 * inch, 2.4 * inch],
                ),
                Spacer(1, 0.18 * inch),
            ]
        )
    if model_outputs or assumptions or gap_warnings:
        story.append(PageBreak())
    generated = payload["content"].get("generated_blocks") or {}
    for index, block in enumerate(payload["content"]["blocks"]):
        if index and index % 3 == 0:
            story.append(PageBreak())
        story.append(Paragraph(html.escape(_block_title(block)), styles["Heading2"]))
        kind = block["kind"]
        if kind in {"NARRATIVE", "LIMITATIONS", "HEADING"}:
            text = str(block.get("text", ""))
            story.append(Paragraph(html.escape(text).replace("\n", "<br/>"), styles["BodyText"]))
        elif kind == "EVIDENCE_REGISTER":
            rows = [["Source", "Blocks", "Claim"]]
            rows.extend(
                [
                    str(citation.get("source_id", "")),
                    ", ".join(citation.get("block_ids") or []),
                    str(citation.get("claim", "")),
                ]
                for citation in block.get("citations") or []
            )
            story.append(Table(rows, repeatRows=1, colWidths=[1.25 * inch, 1.2 * inch, 4.1 * inch]))
        else:
            rows = [["Field", "Frozen value"]]
            rows.extend(
                [str(key), str(value)]
                for key, value in _flatten(generated.get(block["block_id"], {}))
            )
            story.append(Table(rows, repeatRows=1, colWidths=[3.0 * inch, 3.55 * inch]))
        citations = _citation_lines(block)
        for citation in citations:
            story.append(Paragraph(html.escape(citation), styles["CAOSLabel"]))
        story.append(Spacer(1, 0.18 * inch))
    story.extend(
        [
            PageBreak(),
            Paragraph("Revision Record", styles["Heading2"]),
            Table(
                [
                    ["Draft", f"{payload['draft']['id']} v{payload['draft']['version']}"],
                    ["Draft digest", payload["draft"]["digest"]],
                    ["Renderer", payload["renderer"]["version"]],
                    ["Renderer contract", payload["renderer"]["contract_digest"]],
                    ["Preview digest", payload["preview_digest"]],
                ],
                colWidths=[1.55 * inch, 5.1 * inch],
            ),
        ]
    )
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9e7df")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#15151a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("LEADING", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#b5b2aa")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
    )
    for element in story:
        if isinstance(element, Table):
            element.setStyle(table_style)
    document.build(story, canvasmaker=_InvariantCanvas)
    rendered = output.getvalue()
    reader = PdfReader(io.BytesIO(rendered))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if len(reader.pages) < 2 or _title_for(payload) not in text:
        raise ValueError("DELIVERABLE_PDF_SEMANTIC_VERIFICATION_FAILED")
    return rendered


def _safe_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, Decimal)):
        return value
    text = str(value)
    if text.startswith(FORMULA_PREFIXES):
        return "'" + text
    return text


def _style_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="202433")
        cell.alignment = Alignment(vertical="top")
    for column in range(1, sheet.max_column + 1):
        width = max(
            11,
            min(
                55,
                max(
                    len(str(sheet.cell(row=row, column=column).value or ""))
                    for row in range(1, sheet.max_row + 1)
                )
                + 2,
            ),
        )
        sheet.column_dimensions[get_column_letter(column)].width = width


def _append_rows(sheet: Any, rows: Iterable[Iterable[Any]]) -> None:
    for row in rows:
        sheet.append([_safe_cell(value) for value in row])
    _style_sheet(sheet)


_CORE_XML_MODIFIED = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _freeze_zip_timestamps(data: bytes) -> bytes:
    """Rewrite every ZIP entry's date_time and docProps/core.xml's dcterms:modified
    to a fixed value, undoing openpyxl's unconditional wall-clock stamping so the
    frozen export is byte-reproducible for identical input."""
    fixed_iso = _DETERMINISTIC_TIMESTAMP.strftime(r"%Y-%m-%dT%H:%M:%SZ").encode("ascii")
    buffer = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as source:
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                content = source.read(item.filename)
                if item.filename == "docProps/core.xml":
                    content = _CORE_XML_MODIFIED.sub(rb"\g<1>" + fixed_iso + rb"\g<2>", content)
                item.date_time = _DETERMINISTIC_ZIP_DATE_TIME
                target.writestr(item, content)
    return buffer.getvalue()


def render_frozen_xlsx(payload: dict[str, Any]) -> bytes:
    """Render a typed, formula-injection-safe workbook from frozen authority."""

    _validate_payload(payload)
    workbook = Workbook()
    workbook.properties.created = _DETERMINISTIC_TIMESTAMP
    workbook.properties.modified = _DETERMINISTIC_TIMESTAMP
    workbook.remove(workbook.active)
    cover = workbook.create_sheet("Cover")
    _append_rows(
        cover,
        [
            ["Field", "Frozen value"],
            ["Title", _title_for(payload)],
            ["Pathway", payload["pathway"]],
            ["Case ID", payload["case_id"]],
            ["Draft", f"{payload['draft']['id']} v{payload['draft']['version']}"],
            ["Accepted snapshot", payload["authority"].get("accepted_snapshot_id")],
            ["Source set", payload["authority"].get("source_set_id")],
            ["Model build", (payload.get("model") or {}).get("build_id", "Not included")],
            ["Preview digest", payload["preview_digest"]],
        ],
    )
    reviewed = workbook.create_sheet("Reviewed Deliverable")
    reviewed_rows: list[list[Any]] = [["Order", "Kind", "Block ID", "Content"]]
    for order, block in enumerate(payload["content"]["blocks"], start=1):
        reviewed_rows.append(
            [
                order,
                block["kind"],
                block["block_id"],
                block.get("text") or block.get("title") or _block_title(block),
            ]
        )
    _append_rows(reviewed, reviewed_rows)

    summary = workbook.create_sheet("Section Summary")
    _append_rows(
        summary,
        [["Order", "Section", "Kind"]]
        + [
            [index, _block_title(block), block["kind"]]
            for index, block in enumerate(payload["content"]["blocks"], start=1)
        ],
    )

    generated = payload["content"].get("generated_blocks") or {}
    flattened_generated = [
        [block_id, field, value]
        for block_id, value in sorted(generated.items())
        for field, value in _flatten(value)
    ]
    model_outputs = _model_output_rows(payload)
    if flattened_generated or model_outputs:
        detail = workbook.create_sheet("Analytical Detail")
        _append_rows(
            detail,
            [["Block ID", "Field", "Frozen value"]]
            + flattened_generated
            + [
                ["model.outputs", f"{case}.{period_id}.{output_id}", value]
                for case, period_id, output_id, value in model_outputs
            ],
        )
    if model_outputs or any(
        "BASE" in str(row[1]).upper() or "DOWNSIDE" in str(row[1]).upper()
        for row in flattened_generated
    ):
        case_sheet = workbook.create_sheet("Base Downside")
        _append_rows(
            case_sheet,
            [["Case", "Period", "Output", "Frozen value"]]
            + model_outputs
            + [
                [block_id, "", field, value]
                for block_id, field, value in flattened_generated
            ],
        )
    if any(block.get("kind") == "SCENARIO_EXHIBIT" for block in payload["content"]["blocks"]):
        scenario_sheet = workbook.create_sheet("Scenario Analysis")
        scenario_rows = [["Block ID", "Field", "Frozen value"]]
        for block in payload["content"]["blocks"]:
            if block.get("kind") == "SCENARIO_EXHIBIT":
                scenario_rows.extend(
                    [block["block_id"], key, value]
                    for key, value in _flatten(generated.get(block["block_id"], {}))
                )
        _append_rows(scenario_sheet, scenario_rows)
    if payload.get("model"):
        model_sheet = workbook.create_sheet("Model")
        _append_rows(
            model_sheet,
            [["Field", "Frozen value"]]
            + [[key, value] for key, value in _flatten(payload["model"])],
        )
        assumptions = workbook.create_sheet("Assumptions")
        _append_rows(
            assumptions,
            [["Assumption", "Case", "Period", "Unit", "Status", "Value", "Gap"]]
            + _model_assumption_rows(payload),
        )
    debt_rows = _model_debt_rows(payload)
    if debt_rows or any("debt" in str(row[1]).lower() for row in flattened_generated):
        debt = workbook.create_sheet("Debt Schedule")
        _append_rows(
            debt,
            [["Case", "Period", "Output", "Frozen value"]]
            + debt_rows
            + [
                [block_id, "", field, value]
                for block_id, field, value in flattened_generated
                if "debt" in str(field).lower()
            ],
        )
    gaps = _model_gap_warning_rows(payload) + [
        [block_id, field, value]
        for block_id, field, value in flattened_generated
        if value is None
        or "unavailable" in str(value).lower()
        or "warning" in str(field).lower()
        or "gap" in str(field).lower()
    ]
    if gaps:
        gap_sheet = workbook.create_sheet("Gaps and Warnings")
        _append_rows(gap_sheet, [["Block ID", "Field", "Frozen value"]] + gaps)

    evidence = workbook.create_sheet("Evidence Register")
    evidence_rows = [["Source ID", "SHA-256", "Blocks", "Withdrawn"]]
    evidence_rows.extend(
        [
            item.get("source_id"),
            item.get("sha256"),
            ", ".join(item.get("block_ids") or []),
            item.get("withdrawn"),
        ]
        for item in payload.get("evidence") or []
    )
    _append_rows(evidence, evidence_rows)
    revision = workbook.create_sheet("Revision Record")
    _append_rows(
        revision,
        [
            ["Field", "Frozen value"],
            ["Draft ID", payload["draft"]["id"]],
            ["Draft version", payload["draft"]["version"]],
            ["Draft digest", payload["draft"]["digest"]],
            ["Template ID", payload["template"]["template_id"]],
            ["Template version", payload["template"]["template_version"]],
            ["Renderer version", payload["renderer"]["version"]],
            ["Renderer contract", payload["renderer"]["contract_digest"]],
            ["Methodology build", payload["methodology"]["build_id"]],
            ["Input fingerprint", payload["input_fingerprint"]],
            ["Preview digest", payload["preview_digest"]],
        ],
    )
    workbook.properties.title = _title_for(payload)
    workbook.properties.creator = "CAOS"
    output = io.BytesIO()
    workbook.save(output)
    rendered = _freeze_zip_timestamps(output.getvalue())
    reopened = load_workbook(io.BytesIO(rendered), data_only=False, read_only=True)
    if not set(BASE_SHEETS).issubset(reopened.sheetnames):
        raise ValueError("DELIVERABLE_XLSX_SEMANTIC_VERIFICATION_FAILED")
    for sheet in reopened.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError("DELIVERABLE_XLSX_UNTRUSTED_FORMULA")
    return rendered


def render_frozen_exports(payload: dict[str, Any]) -> dict[str, bytes]:
    """Materialize all immutable formats from one exact payload."""

    frozen = deepcopy(payload)
    return {
        "md": render_frozen_markdown(frozen),
        "pdf": render_frozen_pdf(frozen),
        "xlsx": render_frozen_xlsx(frozen),
    }
