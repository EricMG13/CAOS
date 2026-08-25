from __future__ import annotations

import hashlib
import io
import json
import math
import os
import secrets
import socket
import struct
import zipfile
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile

from ..config import Settings
from ..store import MemoryStore, now_iso


ALLOWED_SUFFIXES = {".pdf", ".xlsx", ".json", ".txt", ".md", ".csv"}
MAX_ZIP_ENTRIES = 2000
MAX_ZIP_RATIO = 100
MAX_ZIP_UNCOMPRESSED = 100_000_000
MAX_SOURCE_TEXT = 2_000_000
MAX_SOURCE_LINE = 20_000
MAX_XLSX_EXTRACT_WORKSHEETS = 64
MAX_XLSX_EXTRACT_ROWS = 25_000
MAX_XLSX_EXTRACT_COLUMNS = 64
EXTRACTION_LIMIT_DETAIL = "source exceeds safe extraction limits"


class Vault:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.root = settings.storage_dir
        self.root.mkdir(parents=True, exist_ok=True)

    def put(self, content: bytes, sha256: str) -> str:
        actual = hashlib.sha256(content).hexdigest()
        if actual != sha256:
            raise ValueError("content digest mismatch")
        directory = self.root / "sources" / sha256[:2]
        directory.mkdir(parents=True, exist_ok=True)
        target = directory / sha256
        if target.exists():
            return str(target)
        temp = directory / f".{sha256}.{secrets.token_hex(8)}.tmp"
        try:
            with temp.open("xb") as handle:
                handle.write(content)
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(temp, target)
        finally:
            temp.unlink(missing_ok=True)
        return str(target)


def scan_content(content: bytes, settings: Settings) -> None:
    if b"EICAR-STANDARD-ANTIVIRUS-TEST-FILE" in content:
        raise HTTPException(status_code=422, detail="malware detected")
    if settings.environment != "production":
        return
    if not settings.clamav_host:
        raise HTTPException(status_code=503, detail="malware scanner is not configured")
    try:
        with socket.create_connection((settings.clamav_host, settings.clamav_port), timeout=15) as connection:
            connection.sendall(b"zINSTREAM\0")
            for start in range(0, len(content), 1024 * 1024):
                chunk = content[start : start + 1024 * 1024]
                connection.sendall(struct.pack(">I", len(chunk)) + chunk)
            connection.sendall(b"\0\0\0\0")
            response = connection.recv(4096).lower()
    except OSError as exc:
        raise HTTPException(status_code=503, detail="malware scanner unavailable") from exc
    if b"found" in response:
        raise HTTPException(status_code=422, detail="malware detected")
    if b"ok" not in response:
        raise HTTPException(status_code=503, detail="malware scanner returned an unusable result")


def validate_archive(content: bytes) -> None:
    if not content[:2] == b"PK":
        return
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise HTTPException(status_code=422, detail="archive has too many entries")
            total_size = sum(info.file_size for info in infos)
            if total_size > MAX_ZIP_UNCOMPRESSED or total_size > max(len(content), 1) * MAX_ZIP_RATIO:
                raise HTTPException(status_code=422, detail="archive expanded size exceeds limit")
            for info in infos:
                if info.filename.startswith(("/", "\\")) or ".." in Path(info.filename).parts:
                    raise HTTPException(status_code=422, detail="archive path traversal")
                if info.compress_size and info.file_size / info.compress_size > MAX_ZIP_RATIO:
                    raise HTTPException(status_code=422, detail="archive compression ratio exceeds limit")
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=422, detail="invalid archive") from exc


def extract_blocks(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":
        def reject_non_finite(_constant: str) -> None:
            raise ValueError("non-finite JSON number")

        def reject_overflowing_number(token: str) -> float:
            # `parse_constant` only fires on the NaN/Infinity literals. A finite-looking
            # token that overflows the double range (1e999) is admitted by the parser as
            # `inf`, which then re-serialises to a bare `Infinity` this same reader would
            # reject -- a figure that no longer round-trips through its own evidence.
            value = float(token)
            if not math.isfinite(value):
                raise ValueError("non-finite JSON number")
            return value

        def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
            data = dict(pairs)
            if len(data) != len(pairs):
                raise ValueError("duplicate JSON key")
            return data

        try:
            data = json.loads(content.decode("utf-8"), parse_constant=reject_non_finite, parse_float=reject_overflowing_number, object_pairs_hook=reject_duplicate_keys)
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            raise HTTPException(status_code=422, detail="invalid JSON source") from exc
        if data is None or (isinstance(data, str) and not data.strip()) or (isinstance(data, (dict, list)) and not data):
            raise HTTPException(status_code=422, detail="source contains no extractable evidence")
        text = json.dumps(data, sort_keys=True, indent=2)
    elif suffix == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(content))
            page_count = len(reader.pages)
            if not page_count:
                raise ValueError("PDF contains no pages")
        except Exception as exc:
            raise HTTPException(status_code=422, detail="invalid PDF source") from exc
        try:
            text = "\n".join(reader.pages[index].extract_text() or "" for index in range(page_count))
        except Exception:
            text = ""
        if not text.strip():
            text = "PDF source stored; no text layer was available for safe extraction."
    elif suffix == ".xlsx":
        workbook = None
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
            if len(workbook.worksheets) > MAX_XLSX_EXTRACT_WORKSHEETS:
                raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
            lines: list[str] = []
            text_length = 0

            def append_xlsx_line(line: str) -> None:
                nonlocal text_length
                next_length = text_length + len(line) + (1 if lines else 0)
                if next_length > MAX_SOURCE_TEXT:
                    raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
                lines.append(line)
                text_length = next_length

            has_cell_value = False
            extracted_rows = 0
            for sheet in workbook.worksheets:
                sheet.reset_dimensions()
                append_xlsx_line(f"[sheet:{sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    extracted_rows += 1
                    if extracted_rows > MAX_XLSX_EXTRACT_ROWS:
                        raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
                    if len(row) > MAX_XLSX_EXTRACT_COLUMNS:
                        raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
                    values = ["" if value is None else str(value) for value in row]
                    values.extend([""] * (MAX_XLSX_EXTRACT_COLUMNS - len(values)))
                    has_cell_value = has_cell_value or any(value.strip() for value in values)
                    line = "\t".join(values)
                    append_xlsx_line(line)
            if not has_cell_value:
                raise HTTPException(status_code=422, detail="source contains no extractable evidence")
            text = "\n".join(lines)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=422, detail="invalid XLSX source") from exc
        finally:
            if workbook is not None:
                workbook.close()
    else:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=422, detail="text source must be UTF-8") from exc
    if suffix in {".txt", ".md"} and not text.strip():
        raise HTTPException(status_code=422, detail="source contains no extractable evidence")
    if suffix == ".csv" and not text.strip(" \t\r\n,\""):
        raise HTTPException(status_code=422, detail="source contains no extractable evidence")
    if len(text) > MAX_SOURCE_TEXT:
        raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
    blocks = []
    for index, line in enumerate(text.splitlines() or [text], start=1):
        if line.strip():
            if len(line) > MAX_SOURCE_LINE:
                raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
            blocks.append({"block_id": f"b{index:05d}", "locator": {"line": index}, "text": line, "extractor_version": "builtin-v1", "confidence": "MEDIUM", "untrusted_data": True})
    return blocks or [{"block_id": "b00001", "locator": {"line": 1}, "text": "", "extractor_version": "builtin-v1", "confidence": "LOW", "untrusted_data": True}]


async def ingest_upload(store: MemoryStore, vault: Vault, case_id: str, actor: str, upload: UploadFile, max_bytes: int) -> dict[str, Any]:
    filename = Path(upload.filename or "source.bin").name
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise HTTPException(status_code=415, detail="unsupported source type")
    content = await upload.read(max_bytes + 1)
    if not content:
        raise HTTPException(status_code=422, detail="source is empty")
    if len(content) > max_bytes:
        raise HTTPException(status_code=413, detail="source exceeds upload limit")
    scan_content(content, vault.settings)
    validate_archive(content)
    sha256 = hashlib.sha256(content).hexdigest()
    blocks = extract_blocks(filename, content)
    vault_path = vault.put(content, sha256)
    source_id = store._id("src")
    source = {
        "id": source_id,
        "case_id": case_id,
        "filename": filename,
        "media_type": upload.content_type or "application/octet-stream",
        "bytes": len(content),
        "sha256": sha256,
        "vault_path": vault_path,
        "blocks": blocks,
        "created_by": actor,
        "created_at": now_iso(),
        "withdrawn": False,
    }
    with store.lock:
        if any(existing["case_id"] == case_id and existing["sha256"] == sha256 and not existing.get("withdrawn") for existing in store.sources.values()):
            raise HTTPException(status_code=409, detail="source content already active")
        prior_source_set = store.source_sets.get(case_id)
        audit_start = len(store.audit)
        store.sources[source_id] = source
        source_set = store.source_sets.get(case_id)
        version = (source_set["version"] + 1) if source_set else 1
        active_source_ids = [existing_id for existing_id in (source_set["source_ids"] if source_set else []) if not store.sources.get(existing_id, {}).get("withdrawn")]
        new_source_set = {
            "id": store._id("set"),
            "case_id": case_id,
            "version": version,
            "source_ids": [*active_source_ids, source_id],
            "created_by": actor,
            "created_at": now_iso(),
        }
        store.register_source_set(new_source_set)
        store.audit_event("source.ingested", actor, case_id=case_id, source_id=source_id, sha256=sha256)
        try:
            store.persist()
        except Exception:
            store.sources.pop(source_id, None)
            if prior_source_set is None:
                store.source_sets.pop(case_id, None)
            else:
                store.source_sets[case_id] = prior_source_set
            store.source_set_history.pop(new_source_set["id"], None)
            del store.audit[audit_start:]
            raise
    return {key: value for key, value in {**source, "source_set": store.source_sets[case_id].copy()}.items() if key != "vault_path"}


def list_sources(store: MemoryStore, case_id: str) -> list[dict[str, Any]]:
    with store.lock:
        return [{key: value for key, value in source.items() if key != "vault_path"} for source in store.sources.values() if source["case_id"] == case_id and not source.get("withdrawn")]


def current_source_set(store: MemoryStore, case_id: str) -> dict[str, Any] | None:
    with store.lock:
        source_set = store.source_sets.get(case_id)
        return dict(source_set) if source_set else None


def pathway_fit(store: MemoryStore, case_id: str) -> dict[str, Any]:
    sources = list_sources(store, case_id)
    types = sorted({Path(source["filename"]).suffix.lower().lstrip(".") for source in sources})
    return {
        "source_count": len(sources),
        "file_types": types,
        "fit": "READY" if sources else "NEEDS_SOURCE",
        "language": "Pathway fit is a source-coverage signal, not CP-0 readiness.",
        "message": "Source coverage supports pathway selection." if sources else "Upload governed source material before selecting a route.",
    }
