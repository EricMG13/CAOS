from __future__ import annotations

import hashlib
import math
import os
import secrets
import shutil
import tempfile
import threading
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..artifacts.domain import build_snapshot_payload, latest_accepted_snapshot
from ..contracts import canonical_json, digest
from ..ledgers import ModelLedger, RunLedger, SourceCatalog
from ..methodology.canonical import CANONICAL_MODULES, is_canonical_full_credit
from ..store import JobFencedError
from ..workflows.domain import HEARTBEAT_INTERVAL_SECONDS, _LeaseFence
from .domain import CpModelBundle, ModelInputError


WORKSHEET_SCHEMA_VERSION = "caos.model.worksheet.v1"
VISIBLE_SHEETS = ("Credit Snapshot", "Model", "KPIs")
MAX_WORKSHEET_CELLS = 20_000
MAX_WORKSHEET_BYTES = 8 * 1024 * 1024
MAX_CELL_TEXT = 10_000
MAX_EXPORT_BYTES = 64 * 1024 * 1024


def public_model_build(build: dict[str, Any]) -> dict[str, Any]:
    return {
        key: build.get(key)
        for key in (
            "id",
            "case_id",
            "accepted_run_id",
            "accepted_snapshot_id",
            "source_set_id",
            "input_fingerprint",
            "status",
            "queued_at",
            "started_at",
            "completed_at",
            "error",
            "export",
            "worksheet_schema_version",
            "calculation_runtime",
            "qa",
            "payload_digest",
        )
        if key in build
    }


class ModelReadinessService:
    def __init__(
        self,
        runs: RunLedger,
        sources: SourceCatalog,
        models: ModelLedger,
        methodology: Any,
        model: CpModelBundle,
    ) -> None:
        self.runs = runs
        self.sources = sources
        self.models = models
        self.methodology = methodology
        self.model = model

    def readiness(self, case_id: str) -> dict[str, Any]:
        snapshot = latest_accepted_snapshot(self.runs, case_id)
        if snapshot is None:
            return self._not_ready(
                "ACCEPTED_FULL_CREDIT_REQUIRED",
                "Accept a completed Full Credit run before building a model.",
            )
        try:
            resolved = self._resolve(snapshot)
        except (KeyError, TypeError, ValueError, ModelInputError):
            return self._not_ready(
                "CANONICAL_MODEL_INPUTS_INVALID",
                "The visible accepted snapshot is not a valid canonical Full Credit input.",
            )
        build = next(
            (
                item
                for item in self.models.list_builds(case_id)
                if item.get("input_fingerprint") == resolved["input_fingerprint"]
            ),
            None,
        )
        return {
            "status": build["status"] if build else "READY_TO_BUILD",
            "module_id": "CP-MODEL",
            "accepted_snapshot": resolved["accepted_snapshot"],
            "source_set": resolved["source_set"],
            "requirements": resolved["artifact_inventory"],
            "calculation_runtime": self.model.calculation_runtime,
            "worksheet_schema_version": WORKSHEET_SCHEMA_VERSION,
            "blockers": [],
            "build": public_model_build(build) if build else None,
        }

    def queue(self, case_id: str, actor: str) -> tuple[dict[str, Any], bool]:
        snapshot = latest_accepted_snapshot(self.runs, case_id)
        if snapshot is None:
            raise ValueError("MODEL_NOT_READY")
        try:
            resolved = self._resolve(snapshot)
        except (KeyError, TypeError, ValueError, ModelInputError) as exc:
            raise ValueError("MODEL_NOT_READY") from exc
        build = {
            "case_id": case_id,
            "accepted_run_id": snapshot["run_id"],
            "accepted_snapshot_id": snapshot["id"],
            "source_set_id": snapshot["source_set_id"],
            "input_fingerprint": resolved["input_fingerprint"],
            "accepted_snapshot_digest": snapshot.get("digest"),
            "methodology_build_id": self.methodology.build_id,
            "worksheet_schema_version": WORKSHEET_SCHEMA_VERSION,
            "calculation_runtime": self.model.calculation_runtime,
            "artifact_inventory": resolved["artifact_inventory"],
        }
        record, created = self.models.queue_build(build, actor)
        if not created and record.get("status") == "FAILED":
            return self.models.retry_build(record["id"], actor), True
        return record, created

    def inputs_for_build(self, build: dict[str, Any]) -> dict[str, Any]:
        snapshot = self.runs.get_snapshot(build.get("accepted_snapshot_id"))
        if snapshot is None or snapshot.get("case_id") != build.get("case_id"):
            raise ModelInputError("model snapshot is unavailable")
        resolved = self._resolve(snapshot)
        if (
            resolved["input_fingerprint"] != build.get("input_fingerprint")
            or snapshot.get("run_id") != build.get("accepted_run_id")
            or snapshot.get("source_set_id") != build.get("source_set_id")
        ):
            raise ModelInputError("model inputs changed")
        return resolved

    def _resolve(self, snapshot: dict[str, Any]) -> dict[str, Any]:
        run = self.runs.get_run(snapshot["run_id"])
        if (
            run is None
            or not is_canonical_full_credit(run.get("plan") or {})
            or run.get("accepted_snapshot_id") != snapshot.get("id")
        ):
            raise ModelInputError("accepted canonical Full Credit run required")
        expected = build_snapshot_payload(
            self.runs, self.sources, run, self.methodology
        )
        for key in (
            "case_id",
            "run_id",
            "source_set_id",
            "source_set_version",
            "artifacts",
        ):
            if snapshot.get(key) != expected.get(key):
                raise ModelInputError("accepted snapshot identity mismatch")
        by_module: dict[str, dict[str, Any]] = {}
        for reference in snapshot["artifacts"]:
            if reference["module_id"] not in CANONICAL_MODULES:
                continue
            artifact = self.runs.get_artifact(reference["id"])
            if artifact is None or artifact.get("digest") != reference.get("digest"):
                raise ModelInputError("canonical artifact is unavailable")
            by_module[reference["module_id"]] = artifact
        if set(by_module) != set(CANONICAL_MODULES):
            raise ModelInputError("canonical artifacts are incomplete")
        cp2b = by_module["CP-2A"]["derived"]["CP-2B"]
        markdown = {
            module_id: by_module[module_id]["markdown"]
            for module_id in ("CP-1", "CP-1A", "CP-1B", "CP-2", "CP-2G")
        }
        markdown["CP-2B"] = cp2b["markdown"]
        validation = self.model.validate(
            markdown["CP-1"],
            markdown["CP-1A"],
            markdown["CP-1B"],
            markdown["CP-2"],
            markdown["CP-2B"],
            markdown["CP-2G"],
        )
        if validation.errors:
            raise ModelInputError("CP-MODEL bundle validation failed")
        source_set = self.sources.source_set(snapshot["source_set_id"])
        if source_set is None or source_set.get("case_id") != snapshot.get("case_id"):
            raise ModelInputError("accepted source set is unavailable")
        inventory = [
            {
                "module_id": module_id,
                "artifact_id": by_module[module_id]["id"],
                "digest": by_module[module_id]["digest"],
                "status": "READY",
            }
            for module_id in CANONICAL_MODULES
        ]
        inventory.append(
            {
                "module_id": "CP-2B",
                "artifact_id": by_module["CP-2A"]["id"],
                "digest": cp2b["digest"],
                "derived_from": by_module["CP-2A"]["digest"],
                "status": "READY",
            }
        )
        fingerprint = digest(
            {
                "case_id": snapshot["case_id"],
                "accepted_run_id": snapshot["run_id"],
                "accepted_snapshot_id": snapshot["id"],
                "accepted_snapshot_digest": snapshot.get("digest"),
                "source_set": {
                    "id": source_set["id"],
                    "version": source_set["version"],
                    "digest": digest(source_set),
                },
                "artifacts": [
                    {"module_id": item["module_id"], "digest": item["digest"]}
                    for item in inventory
                ],
                "methodology_build_id": self.methodology.build_id,
                "calculation_runtime": self.model.calculation_runtime,
                "worksheet_schema_version": WORKSHEET_SCHEMA_VERSION,
            }
        )
        return {
            "accepted_snapshot": {
                "id": snapshot["id"],
                "run_id": snapshot["run_id"],
                "digest": snapshot.get("digest"),
            },
            "source_set": {
                "id": source_set["id"],
                "version": source_set["version"],
                "digest": digest(source_set),
            },
            "artifact_inventory": inventory,
            "markdown": markdown,
            "input_fingerprint": fingerprint,
        }

    @staticmethod
    def _not_ready(code: str, detail: str) -> dict[str, Any]:
        return {
            "status": "NOT_READY",
            "module_id": "CP-MODEL",
            "accepted_snapshot": None,
            "source_set": None,
            "requirements": [
                {"module_id": module_id, "status": "MISSING"}
                for module_id in (*CANONICAL_MODULES, "CP-2B")
            ],
            "calculation_runtime": None,
            "worksheet_schema_version": WORKSHEET_SCHEMA_VERSION,
            "blockers": [{"code": code, "detail": detail}],
            "build": None,
        }


class ModelBuildRuntime:
    def __init__(
        self,
        models: ModelLedger,
        readiness: ModelReadinessService,
        model: CpModelBundle,
        executor: Any,
        storage_dir: Path,
    ) -> None:
        self.models = models
        self.readiness = readiness
        self.model = model
        self.executor = executor
        self.storage_dir = storage_dir

    def schedule(self, build_id: str, actor: str) -> Any:
        return self.executor.submit(self._execute, build_id, actor)

    def schedule_export(self, build_id: str, actor: str) -> Any:
        return self.executor.submit(self._execute_export, build_id, actor)

    def _execute(self, build_id: str, actor: str) -> None:
        token = self.models.claim(build_id, threading.current_thread().name)
        if token is None:
            return
        heartbeat_stop = threading.Event()
        fence = _LeaseFence()

        def heartbeat() -> None:
            while not heartbeat_stop.wait(HEARTBEAT_INTERVAL_SECONDS):
                try:
                    if self.models.renew(build_id, token):
                        continue
                except Exception:
                    pass
                fence.lose()
                return

        thread = threading.Thread(
            target=heartbeat, name=f"{build_id}-heartbeat", daemon=True
        )
        thread.start()
        try:
            build = self.models.get_build(build_id)
            if build is None:
                raise ModelInputError("model build is unavailable")
            resolved = self.readiness.inputs_for_build(build)
            with tempfile.TemporaryDirectory(prefix="caos-model-") as temporary:
                directory = Path(temporary)
                paths = self._write_inputs(directory, resolved["markdown"])
                model, calculations = self.model.calculate(paths)
                draft = directory / "model.xlsx"
                rendered = self.model.render_workbook(model, calculations, draft)
                payload, qa = _serialize_worksheet(draft, rendered, model, calculations)
            if fence.lost.is_set() or not self.models.is_current(build_id, token):
                raise JobFencedError("lost model lease")
            fence.call(
                self.models.complete,
                build_id,
                token,
                {
                    "payload": payload,
                    "payload_digest": digest(payload),
                    "qa": qa,
                },
                actor,
            )
        except JobFencedError:
            pass
        except Exception as exc:
            error = {
                "code": "MODEL_INPUT_INVALID"
                if isinstance(exc, ModelInputError)
                else "MODEL_CALCULATION_FAILED",
                "detail": "The accepted model inputs are invalid."
                if isinstance(exc, ModelInputError)
                else "The model calculation did not complete.",
            }
            try:
                fence.call(self.models.fail, build_id, token, error, actor)
            except JobFencedError:
                pass
        finally:
            heartbeat_stop.set()
            thread.join(timeout=1)

    def _execute_export(self, build_id: str, actor: str) -> None:
        token = self.models.claim(build_id, threading.current_thread().name, "export")
        if token is None:
            return
        heartbeat_stop = threading.Event()
        fence = _LeaseFence()

        def heartbeat() -> None:
            while not heartbeat_stop.wait(HEARTBEAT_INTERVAL_SECONDS):
                try:
                    if self.models.renew(build_id, token, "export"):
                        continue
                except Exception:
                    pass
                fence.lose()
                return

        thread = threading.Thread(
            target=heartbeat, name=f"{build_id}-export-heartbeat", daemon=True
        )
        thread.start()
        published: Path | None = None
        try:
            build = self.models.get_build(build_id)
            if build is None or build.get("status") != "READY":
                raise ModelInputError("ready model build is unavailable")
            resolved = self.readiness.inputs_for_build(build)
            with tempfile.TemporaryDirectory(prefix="caos-model-export-") as temporary:
                directory = Path(temporary)
                paths = self._write_inputs(directory, resolved["markdown"])
                result = self.model.export(paths, directory / "output")
                published, vault_key, size = _publish_export(
                    self.storage_dir, build, result.output, result.sha256
                )
            if fence.lost.is_set() or not self.models.is_current(
                build_id, token, "export"
            ):
                raise JobFencedError("lost model export lease")
            fence.call(
                self.models.complete,
                build_id,
                token,
                {
                    "vault_key": vault_key,
                    "filename": result.output.name,
                    "sha256": result.sha256,
                    "size": size,
                    "formulas_validated": result.formulas_validated,
                    "semantic_checks": result.semantic_checks,
                    "renderer_version": result.renderer_version,
                    "renderer_sha256": result.renderer_sha256,
                    "calculation_engine": result.calculation_engine[:200],
                },
                actor,
                "export",
            )
        except JobFencedError:
            pass
        except Exception:
            try:
                fence.call(
                    self.models.fail,
                    build_id,
                    token,
                    {
                        "code": "MODEL_EXPORT_FAILED",
                        "detail": "The XLSX export did not complete.",
                    },
                    actor,
                    "export",
                )
            except JobFencedError:
                pass
            else:
                if published is not None:
                    published.unlink(missing_ok=True)
        finally:
            heartbeat_stop.set()
            thread.join(timeout=1)

    @staticmethod
    def _write_inputs(directory: Path, markdown: dict[str, str]) -> dict[str, Path]:
        paths: dict[str, Path] = {}
        for module_id, content in markdown.items():
            path = directory / f"{module_id.lower()}.md"
            path.write_text(content, encoding="utf-8")
            paths[module_id] = path
        return paths


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise ValueError("non-finite worksheet value")
        converted = float(value)
        if not math.isfinite(converted):
            raise ValueError("non-finite worksheet value")
        return converted
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("non-finite worksheet value")
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, str) and len(value) > MAX_CELL_TEXT:
        raise ValueError("worksheet cell exceeds size limit")
    if value is None or isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _publish_export(
    storage_dir: Path,
    build: dict[str, Any],
    source: Path,
    expected_sha256: str,
) -> tuple[Path, str, int]:
    if any(
        not isinstance(value, str)
        or not value
        or any(
            character
            not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_-"
            for character in value
        )
        for value in (build.get("case_id"), build.get("id"))
    ):
        raise ValueError("unsafe model export identity")
    size = source.stat().st_size
    if size <= 0 or size > MAX_EXPORT_BYTES:
        raise ValueError("model export size limit exceeded")
    actual_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
    if actual_sha256 != expected_sha256:
        raise ValueError("model export digest mismatch")
    root = storage_dir.resolve()
    root.mkdir(parents=True, exist_ok=True)
    directory = root / "models" / build["case_id"] / build["id"]
    directory.mkdir(parents=True, exist_ok=True)
    directory = directory.resolve()
    if not directory.is_relative_to(root):
        raise ValueError("model export path escapes storage root")
    target = directory / f"{expected_sha256}.xlsx"
    if target.exists():
        if target.is_symlink() or target.stat().st_size != size:
            raise ValueError("existing model export identity mismatch")
        if hashlib.sha256(target.read_bytes()).hexdigest() != expected_sha256:
            raise ValueError("existing model export digest mismatch")
        return target, target.relative_to(root).as_posix(), size
    temporary = directory / f".{expected_sha256}.{secrets.token_hex(8)}.tmp"
    descriptor: int | None = None
    try:
        descriptor = os.open(temporary, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        with os.fdopen(descriptor, "wb") as output, source.open("rb") as input_file:
            descriptor = None
            shutil.copyfileobj(input_file, output)
            output.flush()
            os.fsync(output.fileno())
        os.replace(temporary, target)
    finally:
        if descriptor is not None:
            os.close(descriptor)
        temporary.unlink(missing_ok=True)
    return target, target.relative_to(root).as_posix(), size


def _cell_type(value: Any, formula: str | None) -> str:
    if formula:
        return "formula"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    if isinstance(value, (date, datetime)):
        return "date"
    return "text"


def _rgb(color: Any) -> str | None:
    value = getattr(color, "rgb", None)
    if getattr(color, "type", None) != "rgb" or not isinstance(value, str):
        return None
    return value[-6:]


def _serialize_worksheet(
    draft: Path,
    rendered: Any,
    model: Any,
    calculations: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    formulas = {(item.sheet, item.cell): item for item in rendered.formulas}
    mappings = {item.target: item for item in rendered.mappings}
    workbook = load_workbook(draft, data_only=False, read_only=False)
    try:
        tabs: list[dict[str, Any]] = []
        cell_count = 0
        for sheet_name in VISIBLE_SHEETS:
            worksheet = workbook[sheet_name]
            cells: list[dict[str, Any]] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    address = cell.coordinate
                    expectation = formulas.get((sheet_name, address))
                    if cell.value is None and expectation is None:
                        continue
                    formula = (
                        cell.value
                        if isinstance(cell.value, str) and cell.value.startswith("=")
                        else None
                    )
                    value = (
                        expectation.expected if expectation is not None else cell.value
                    )
                    mapping = mappings.get(f"{sheet_name}!{address}")
                    cells.append(
                        {
                            "address": address,
                            "row": cell.row,
                            "column": cell.column,
                            "value": _json_value(value),
                            "value_type": _cell_type(value, formula),
                            "formula": formula,
                            "semantic_id": mapping.semantic_id if mapping else None,
                            "owner": mapping.owner if mapping else None,
                            "write_class": mapping.write_class if mapping else None,
                            "period_id": mapping.period_id if mapping else None,
                            "source_refs": mapping.source_refs if mapping else None,
                            "number_format": cell.number_format,
                            "style": {
                                "bold": bool(cell.font.bold),
                                "italic": bool(cell.font.italic),
                                "fill": _rgb(cell.fill.fgColor),
                                "align": cell.alignment.horizontal,
                                "wrap": bool(cell.alignment.wrap_text),
                            },
                        }
                    )
                    cell_count += 1
                    if cell_count > MAX_WORKSHEET_CELLS:
                        raise ValueError("worksheet cell limit exceeded")
            columns = []
            for index in range(1, worksheet.max_column + 1):
                letter = get_column_letter(index)
                dimension = worksheet.column_dimensions[letter]
                columns.append(
                    {
                        "column": index,
                        "letter": letter,
                        "width": dimension.width,
                        "hidden": bool(dimension.hidden),
                    }
                )
            tabs.append(
                {
                    "id": sheet_name.upper().replace(" ", "_"),
                    "name": sheet_name,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "freeze_panes": str(worksheet.freeze_panes or ""),
                    "merged_cells": [
                        str(item) for item in worksheet.merged_cells.ranges
                    ],
                    "columns": columns,
                    "cells": cells,
                }
            )
        payload = {
            "schema_version": WORKSHEET_SCHEMA_VERSION,
            "identity": {
                "issuer_id": model.issuer_id,
                "issuer_name": model.issuer_name,
                "analysis_date": model.analysis_date.isoformat(),
            },
            "tabs": tabs,
        }
        if len(canonical_json(payload).encode("utf-8")) > MAX_WORKSHEET_BYTES:
            raise ValueError("worksheet payload limit exceeded")
        checks = [
            {
                "check_id": check.check_id,
                "status": check.status,
                "period_id": check.period_id,
                "difference": _json_value(check.difference),
                "tolerance": _json_value(check.tolerance),
                "detail": check.detail,
            }
            for check in calculations.checks
        ]
        qa = {
            "status": "PASS",
            "semantic_checks": checks,
            "semantic_check_count": len(checks),
            "formula_count": len(rendered.formulas),
            "worksheet_cell_count": cell_count,
            "limitation_flags": list(model.limitation_flags),
            "validation_warnings": list(model.validation_warnings),
            "source_manifest": [
                {
                    "module_id": module_id,
                    "filename": model.source_paths[module_id].name,
                    "sha256": model.source_hashes[module_id],
                }
                for module_id in sorted(model.source_paths)
            ],
        }
        return payload, qa
    finally:
        workbook.close()
